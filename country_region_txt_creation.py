"""
CountryZoning TXT Generator
---------------------------
Builds a plain-text summary from CountryZoning data (preferably from
``extracted_data.json``).

**Forming zones (Export / Import)**

1. Find header cells such as ``Service1``: ``Zonenumber Sending`` and
   ``Service8``: ``Zonenumber Receiving``. The **smallest** ``ServiceN`` index
   whose cell contains *receiving* is the split: columns below that index are
   **Export** (Sending); that index and above are **Import** (Receiving).
   If only Sending is found, **8** is used as the default split (Service1–7
   Export, Service8+ Import).

2. Product names come from (a) early table header rows and (b) **embedded legend
   rows** in the country grid, e.g. ``Country: Lithuania`` with
   ``Service1``: ``express plus\\nZonenumber Sending``, ``Service2``: ``express\\n41``.
   Parse lines until ``Zonenumber``/``Zonnummer``; skip chart digits (``41``, …).

3. Zone lines: ``{Product} {Export|Import} Zone {n} - countries``
   Section line: ``Service1: express plus export`` (product lowercased +
   export/import).

**Country Regions TXT — primary rows**

- Aggregate by ``Code`` (ISO-style), not ``Country`` display names.
- Include only when: no ``()`` in CustomerCountry; empty PostCode1/PostCode2;
  and ``Country`` does **not** contain ``excluding`` (case-insensitive).

**Postal Code Zones TXT — separate file**

- Rows excluded from Country Regions: ``excluding`` in ``Country`` (Case 1),
  parenthetical CustomerCountry, and/or non-empty postcodes.
- Case 1: ``country (excluding A and B)`` plus child rows ``A``, ``B`` with the
  same customer/code — structured columns: zone name, country code, postal,
  excluded.
- Case 2: same ``Code`` but different ``Country`` (e.g. GB + England/Scotland/
  Wales/Northern Ireland). Country Regions gets ``Code`` only when **every**
  subdivision row has that ``Service*`` with the **same** zone; otherwise detail
  goes here with ``GB (subregions)`` zone titles.
- Case 3: postcodes in **Customer country** text, e.g.
  ``Sweden (20000-23799)``, ``Sweden (other postcode)`` — one postal line per
  range and an ``(other)`` line with ``0-9`` and excluded set to the explicit
  ranges. (Skipped when ``PostCode1``/``PostCode2`` are set — Case 4.)
- Case 4: postcodes in **PostCode1** / **PostCode2**. Group by **Code** (ignore
  ``CustomerCountry``). Rows with the same service-zone **signature** merge
  numeric ranges; ``other``/``övriga`` rows use ``0-9`` and excluded = prior
  numeric ranges. Zone names: ``Export express plus DK 2`` (``{EI} {product}
  {CODE} {zone}``).

Optional trailing sections (Country Regions file only): DemandSurchargeCountries,
GoGreenPlusCost.
"""

import json
import re
from collections import defaultdict
from pathlib import Path

from transform_other_tabs import (
    _load_country_codes,
    _gogreen_country_list_to_codes,
    build_gogreen_block_txt_lines,
    demand_surcharge_origin_label,
    demand_surcharge_destination_label,
)

_SERVICE_KEY = re.compile(r"^Service\d+$", re.I)


def _str(v):
    if v is None:
        return ""
    return str(v).strip()


def _nonblank(v):
    return v is not None and str(v).strip() != ""


def _is_zone_value(s: str) -> bool:
    """True if the cell looks like a zone index (integer string, e.g. 9, 61)."""
    if not s:
        return False
    s = s.strip()
    return bool(re.match(r"^-?\d+$", s))


def _pretty_service_title(phrase: str) -> str:
    """Title-case a service / product phrase; keep common carrier acronyms."""
    if not phrase:
        return phrase
    t = " ".join(phrase.split()).title()
    fixes = {"Ups": "UPS", "Dhl": "DHL", "Us": "US", "Uk": "UK", "Eu": "EU"}
    return " ".join(fixes.get(w, w) for w in t.split())


# Default split: Service1–7 Export (Sending), Service8+ Import (Receiving) when
# the Zonenumber Receiving column cannot be detected.
_DEFAULT_EXPORT_IMPORT_SPLIT = 8

# Standard UPS CountryZoning column layout (Service1–14) for Postal Code Zones
# zone titles when headers do not supply product names.
_DEFAULT_UPS_SERVICE_PRODUCT: dict[int, str] = {
    1: "express plus",
    2: "express",
    3: "express saver",
    4: "standard",
    5: "expedited",
    6: "express freight",
    7: "express freight Midday",
    8: "express plus",
    9: "express",
    10: "express saver",
    11: "standard",
    12: "expedited",
    13: "express freight",
    14: "express freight Midday",
}

# Mainland postal band for Case 1 parent rows (excluding … in Country) when
# no finer-grained data exists in the extract.
_CASE1_MAINLAND_POSTAL_PLACEHOLDER = "0-9"


def _country_has_excluding(country: str) -> bool:
    return "excluding" in (country or "").lower()


def _parse_excluded_regions_from_country(country: str) -> list[str]:
    """
    From ``Portugal (excluding Azores and Madeira)`` return
    ``['Azores', 'Madeira']``.
    """
    s = (country or "").strip()
    if not s:
        return []
    m = re.search(r"\(\s*excluding\s+(.+?)\)\s*$", s, re.I | re.DOTALL)
    if not m:
        m = re.search(r"\bexcluding\s+(.+)$", s, re.I | re.DOTALL)
    if not m:
        return []
    inner = m.group(1).strip().rstrip(")").strip()
    parts = re.split(r"\s+and\s+|\s*,\s*", inner)
    return [p.strip() for p in parts if p.strip()]


def _norm_key(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def _postal_product_phrase(
    service_key: str, column_products: dict[str, str]
) -> str:
    n = _service_column_number(service_key)
    if n in _DEFAULT_UPS_SERVICE_PRODUCT:
        return _DEFAULT_UPS_SERVICE_PRODUCT[n]
    raw = (column_products.get(service_key) or "").strip()
    if raw:
        return " ".join(raw.split()).lower()
    return f"service {n}"


def _postal_zone_title(
    service_key: str,
    zone: str,
    code: str,
    split_idx: int | None,
    column_products: dict[str, str],
    subregion: str | None,
) -> str:
    """
    e.g. ``Export express plus 4 PT`` or ``Export express saver 4 PT(Azores)``.
    """
    col_num = _service_column_number(service_key)
    ei = _export_or_import(col_num, split_idx)
    prod = _postal_product_phrase(service_key, column_products)
    c = (code or "").strip().upper()
    if subregion:
        reg = subregion.strip()
        return f"{ei} {prod} {zone} {c}({reg})"
    return f"{ei} {prod} {zone} {c}"


def _service_column_number(k: str) -> int:
    m = re.match(r"(?i)^service(\d+)$", k)
    return int(m.group(1)) if m else 0


def _find_export_import_split(all_rows: list) -> int | None:
    """
    Return N such that ServiceN is the first **Import** (Receiving) column;
    Service1 … Service(N-1) are **Export** (Sending).

    Scans **all** CountryZoning rows (table headers and embedded legend rows
    e.g. Lithuania with ``express plus\\nZonenumber Sending``).
    """
    receiving = []
    for row in all_rows:
        if not isinstance(row, dict):
            continue
        for k, v in row.items():
            if not _SERVICE_KEY.match(k):
                continue
            s = _str(v).lower()
            if "zonenumber" not in s and "zonnummer" not in s:
                continue
            n = _service_column_number(k)
            if not n:
                continue
            if "receiving" in s:
                receiving.append(n)
    if receiving:
        return min(receiving)
    return None


def _parse_product_from_service_cell(raw: str) -> str | None:
    """
    Product name from a ``ServiceN`` cell: lines before ``Zonenumber`` /
    ``Zonnummer``; skip pure-digit chart lines (``41``, ``2``).

    Examples::

        express plus\\nZonenumber Sending  -> express plus
        express\\n41                       -> express
        express plus\\n41\\nZonenumber Receiving -> express plus
        expedited                          -> expedited
    """
    if not raw or not str(raw).strip():
        return None
    lines = [x.strip() for x in str(raw).splitlines() if x.strip()]
    if not lines:
        return None
    if "customer" in lines[0].lower() and "country" in lines[0].lower():
        return None

    parts: list[str] = []
    for line in lines:
        low = line.lower()
        if "zonenumber" in low or "zonnummer" in low:
            break
        if re.match(r"^-?\d+$", line):
            continue
        parts.append(line)

    joined = " ".join(parts).strip()
    if not joined:
        return None
    low = joined.lower()
    if "zonenumber" in low or "zonnummer" in low:
        return None
    return joined


def _find_column_product_names(all_rows: list, header_rows: list) -> dict[str, str]:
    """
    ``ServiceN`` -> product string from table headers **and** embedded legend
    rows (e.g. Country=Lithuania with ``express plus\\nZonenumber Sending``).
    """
    legend_rows = [r for r in all_rows if _is_service_legend_row(r)]
    source = list(header_rows) + legend_rows

    out: dict[str, str] = {}
    scores: dict[str, int] = {}
    for row in source:
        if not isinstance(row, dict):
            continue
        for k, v in row.items():
            if not _SERVICE_KEY.match(k):
                continue
            raw = str(v) if v is not None else ""
            p = _parse_product_from_service_cell(raw)
            if not p:
                continue
            score = len(raw) + raw.count("\n") * 40
            if _is_service_legend_row(row):
                score += 10_000
            if score > scores.get(k, -1):
                out[k] = p
                scores[k] = score
    return out


def _export_or_import(col_num: int, split_idx: int | None) -> str:
    s = split_idx if split_idx is not None else _DEFAULT_EXPORT_IMPORT_SPLIT
    return "Export" if col_num < s else "Import"


def _is_table_header_template(row: dict) -> bool:
    cc = _str(row.get("CustomerCountry"))
    if "Customer" in cc and "Code" in cc:
        return True
    return False


def _has_country_field(row: dict) -> bool:
    return _nonblank(row.get("Country"))


def _is_data_country_row(row: dict) -> bool:
    """A grid row: has Country and at least one numeric Service* zone."""
    if _is_table_header_template(row):
        return False
    if not _has_country_field(row):
        return False
    c = _str(row.get("Country"))
    if c.upper() == "COUNTRY":
        return False
    for k, v in row.items():
        if not _SERVICE_KEY.match(k):
            continue
        if _is_zone_value(_str(v)):
            return True
    return False


def _first_data_row_index(rows: list) -> int:
    for i, row in enumerate(rows):
        if not isinstance(row, dict):
            continue
        if _is_data_country_row(row):
            return i
    return len(rows)


def _is_primary_zoning_row(row: dict) -> bool:
    """Eligible for Country Regions TXT: codes list; excludes postal-only patterns."""
    cc = _str(row.get("CustomerCountry"))
    if "(" in cc or ")" in cc:
        return False
    if _nonblank(row.get("PostCode1")) or _nonblank(row.get("PostCode2")):
        return False
    if _country_has_excluding(_str(row.get("Country"))):
        return False
    return True


def _child_matches_excluded_token(country_display: str, excluded_tokens: list[str]) -> bool:
    nk = _norm_key(country_display)
    for ex in excluded_tokens:
        if nk == _norm_key(ex):
            return True
    return False


def _case1_child_indices(rows: list) -> set[int]:
    """
    Row indices that are Case 1 **subregion** rows (e.g. Azores) linked to a
    parent ``… (excluding …)``. Those rows must not appear in Country Regions TXT.
    """
    out: set[int] = set()
    by_cc_code: dict[tuple[str, str], list[tuple[int, dict]]] = defaultdict(list)
    for i, row in enumerate(rows):
        if not isinstance(row, dict) or not _is_data_country_row(row):
            continue
        cc = _norm_key(_str(row.get("CustomerCountry")))
        code = _str(row.get("Code") or row.get("Country Code")).strip().upper()
        if not code:
            continue
        by_cc_code[(cc, code)].append((i, row))

    for i, row in enumerate(rows):
        if not isinstance(row, dict) or not _is_data_country_row(row):
            continue
        if not _country_has_excluding(_str(row.get("Country"))):
            continue
        excluded = _parse_excluded_regions_from_country(_str(row.get("Country")))
        if not excluded:
            continue
        code = _str(row.get("Code") or row.get("Country Code")).strip().upper()
        cc_key = _norm_key(_str(row.get("CustomerCountry")))
        if not code:
            continue
        for j, r2 in by_cc_code.get((cc_key, code), []):
            if j == i:
                continue
            c2 = _str(r2.get("Country"))
            if _country_has_excluding(c2):
                continue
            if _child_matches_excluded_token(c2, excluded):
                out.add(j)
    return out


def _case2_cluster_key_from_row(row: dict) -> tuple[str, str] | None:
    """``(CustomerCountry, Code)`` for rows that can participate in Case 2."""
    if not isinstance(row, dict) or not _is_data_country_row(row):
        return None
    if not _is_primary_zoning_row(row):
        return None
    code = _str(row.get("Code") or row.get("Country Code")).strip().upper()
    if not code:
        return None
    return (_norm_key(_str(row.get("CustomerCountry"))), code)


def _case2_clusters_map(rows: list) -> dict[tuple[str, str], list[int]]:
    """
    Groups with the same customer + ISO code but **≥2 distinct** ``Country``
    values (Case 2).
    """
    temp: dict[tuple[str, str], list[int]] = defaultdict(list)
    for i, row in enumerate(rows):
        if not isinstance(row, dict):
            continue
        ck = _case2_cluster_key_from_row(row)
        if ck is None:
            continue
        temp[ck].append(i)
    out: dict[tuple[str, str], list[int]] = {}
    for key, inds in temp.items():
        countries = {_str(rows[j].get("Country")) for j in inds if j < len(rows)}
        countries.discard("")
        if len(countries) >= 2:
            out[key] = inds
    return out


def _case2_cluster_country_names(rows: list, cluster_inds: list[int]) -> list[str]:
    names = []
    seen: set[str] = set()
    for j in cluster_inds:
        if j >= len(rows):
            continue
        n = _str(rows[j].get("Country")).strip()
        if not n:
            continue
        nk = _norm_key(n)
        if nk in seen:
            continue
        seen.add(nk)
        names.append(n)
    return sorted(names, key=lambda x: x.lower())


def _case2_unanimous(
    cluster_inds: list[int], rows: list, service_key: str, zone_str: str
) -> bool:
    """True iff every row in the cluster has ``service_key`` = ``zone_str``."""
    if not cluster_inds:
        return False
    for i in cluster_inds:
        r = rows[i]
        v = r.get(service_key)
        if v is None or not _is_zone_value(_str(v)):
            return False
        if _str(v) != zone_str:
            return False
    return True


def _case2_use_bare_gb_title(
    subset_country_names: list[str], all_cluster_country_names: list[str]
) -> bool:
    """
    ``Export standard 703 GB`` without ``(…)`` when the subset is all cluster
    members except **Northern Ireland** only.
    """
    cluster_norm = {_norm_key(c) for c in all_cluster_country_names}
    subset_norm = {_norm_key(c) for c in subset_country_names}
    excluded = cluster_norm - subset_norm
    if len(excluded) != 1:
        return False
    (ex,) = tuple(excluded)
    return ex == "northern ireland" and len(subset_norm) >= 2


def _case2_zone_title(
    service_key: str,
    zone: str,
    code: str,
    split_idx: int | None,
    column_products: dict[str, str],
    subset_country_names: list[str],
    all_cluster_country_names: list[str],
) -> str:
    """
    e.g. ``Export express plus 703 GB (England)`` or bare ``Export standard 703 GB``.
    """
    if _case2_use_bare_gb_title(subset_country_names, all_cluster_country_names):
        col_num = _service_column_number(service_key)
        ei = _export_or_import(col_num, split_idx)
        prod = _postal_product_phrase(service_key, column_products)
        c = code.strip().upper()
        return f"{ei} {prod} {zone} {c}"
    sorted_names = sorted(subset_country_names, key=lambda x: x.lower())
    col_num = _service_column_number(service_key)
    ei = _export_or_import(col_num, split_idx)
    prod = _postal_product_phrase(service_key, column_products)
    c = code.strip().upper()
    if len(sorted_names) == 1:
        return f"{ei} {prod} {zone} {c} ({sorted_names[0]})"
    inner = ", ".join(sorted_names)
    return f"{ei} {prod} {zone} {c} ({inner})"


def _case2_postal_and_excluded_columns(
    subset_country_names: list[str],
    all_cluster_country_names: list[str],
    bare_title: bool,
) -> tuple[str, str]:
    """``(postal column, excluded column)`` for Case 2 postal file lines."""
    cluster_norm_to_display = {
        _norm_key(c): c.strip() for c in all_cluster_country_names if c.strip()
    }
    subset_norm = {_norm_key(c) for c in subset_country_names}
    excluded_norm = set(cluster_norm_to_display.keys()) - subset_norm
    excluded_display = sorted(
        [cluster_norm_to_display[k] for k in excluded_norm if k in cluster_norm_to_display],
        key=lambda x: x.lower(),
    )
    exc_str = ", ".join(excluded_display)

    if bare_title:
        return _CASE1_MAINLAND_POSTAL_PLACEHOLDER, exc_str

    if (
        len(subset_country_names) == 1
        and _norm_key(subset_country_names[0]) == "northern ireland"
    ):
        return "Northern Ireland", exc_str

    return _CASE1_MAINLAND_POSTAL_PLACEHOLDER, exc_str


_CASE3_NUMERIC_RANGE = re.compile(r"^\s*(\d+)\s*-\s*(\d+)\s*$")


def _parse_case3_customer_country(cc: str) -> tuple[str, str] | None:
    """``CustomerCountry`` with a single trailing ``(…)`` → ``(base, inner)``."""
    s = (cc or "").strip()
    if "(" not in s or ")" not in s:
        return None
    m = re.match(r"^\s*(.+?)\s*\(\s*([^)]*)\s*\)\s*$", s)
    if not m:
        return None
    base = m.group(1).strip()
    inner = (m.group(2) or "").strip()
    if not base:
        return None
    return base, inner


def _case3_inner_numeric_range(inner: str) -> str | None:
    m = _CASE3_NUMERIC_RANGE.match(inner.strip())
    if not m:
        return None
    a, b = int(m.group(1)), int(m.group(2))
    return f"{a}-{b}"


def _case3_inner_is_other(inner: str) -> bool:
    low = inner.lower().strip()
    if not low:
        return False
    if low in ("other", "other postcode"):
        return True
    return low.startswith("other") and "postcode" in low


def _is_case3_eligible_row(row: dict) -> bool:
    """Postcode-in-Customer-country pattern (numeric range or ``other``)."""
    if not _is_data_country_row(row):
        return False
    if _nonblank(row.get("PostCode1")) or _nonblank(row.get("PostCode2")):
        return False
    code = _str(row.get("Code") or row.get("Country Code")).strip()
    if not code:
        return False
    parsed = _parse_case3_customer_country(_str(row.get("CustomerCountry")))
    if not parsed:
        return False
    _base, inner = parsed
    if _case3_inner_is_other(inner):
        return True
    if _case3_inner_numeric_range(inner):
        return True
    return False


def _case3_groups(rows: list) -> dict[tuple[str, str], list[int]]:
    """``(norm_customer_label, Code)`` → row indices for Case 3."""
    g: dict[tuple[str, str], list[int]] = defaultdict(list)
    for i, row in enumerate(rows):
        if not isinstance(row, dict) or not _is_case3_eligible_row(row):
            continue
        parsed = _parse_case3_customer_country(_str(row.get("CustomerCountry")))
        code = _str(row.get("Code") or row.get("Country Code")).strip().upper()
        base = parsed[0]
        g[(_norm_key(base), code)].append(i)
    return dict(g)


_DIGITS_ONLY_POSTCODE = re.compile(r"^\d+$")


def _postcode_cells_numeric_pair(p1: str, p2: str) -> bool:
    a, b = _str(p1).strip(), _str(p2).strip()
    return bool(_DIGITS_ONLY_POSTCODE.match(a) and _DIGITS_ONLY_POSTCODE.match(b))


def _format_postcode_range_numeric(p1: str, p2: str) -> str:
    a, b = int(_str(p1).strip()), int(_str(p2).strip())
    lo, hi = (a, b) if a <= b else (b, a)
    return f"{lo}-{hi}"


def _postcode_row_is_other_cells(p1: str, p2: str) -> bool:
    """Non-numeric pair (``other``, ``övriga``, or any non-digit postcode cells)."""
    if _postcode_cells_numeric_pair(p1, p2):
        return False
    s1, s2 = _str(p1).lower(), _str(p2).lower()
    if "other" in s1 or "other" in s2:
        return True
    if "övriga" in s1 or "övriga" in s2 or "ovriga" in s1 or "ovriga" in s2:
        return True
    return True


def _case4_service_signature(row: dict) -> tuple[tuple[str, str], ...]:
    items: list[tuple[str, str]] = []
    for k in sorted([x for x in row if _SERVICE_KEY.match(x)], key=_service_key_sort):
        v = row.get(k)
        if v is not None and _is_zone_value(_str(v)):
            items.append((k, _str(v).strip()))
    return tuple(items)


def _case4_zone_title(
    service_key: str,
    zone: str,
    code: str,
    split_idx: int | None,
    column_products: dict[str, str],
) -> str:
    """``Export express plus DK 2`` — product order: ``{EI} {product} {CODE} {zone}``."""
    col_num = _service_column_number(service_key)
    ei = _export_or_import(col_num, split_idx)
    prod = _postal_product_phrase(service_key, column_products)
    c = code.strip().upper()
    return f"{ei} {prod} {c} {zone}"


def _is_case4_eligible_row(row: dict) -> bool:
    if not _is_data_country_row(row):
        return False
    if not (_nonblank(row.get("PostCode1")) or _nonblank(row.get("PostCode2"))):
        return False
    code = _str(row.get("Code") or row.get("Country Code")).strip()
    return bool(code)


def _case4_indices_by_code(rows: list) -> dict[str, list[int]]:
    d: dict[str, list[int]] = defaultdict(list)
    for i, row in enumerate(rows):
        if not isinstance(row, dict) or not _is_case4_eligible_row(row):
            continue
        code = _str(row.get("Code") or row.get("Country Code")).strip().upper()
        d[code].append(i)
    for c in d:
        d[c].sort()
    return dict(d)


def _is_service_legend_row(row: dict) -> bool:
    """
    True for embedded **service-name** rows (e.g. Lithuania) where ``ServiceN``
    cells hold ``express plus\\nZonenumber Sending`` instead of a numeric zone.

    Such rows are not zone-aggregation rows (no pure integer in any Service cell).
    Exception/postcode rows are excluded so Denmark definition cells are not used.
    """
    if not isinstance(row, dict) or not _has_country_field(row):
        return False
    if _is_table_header_template(row):
        return False
    if not _is_primary_zoning_row(row):
        return False
    svc_keys = [k for k in row if _SERVICE_KEY.match(k)]
    vals = [_str(row[k]) for k in svc_keys if _nonblank(row.get(k))]
    if len(vals) < 3:
        return False
    if any(_is_zone_value(v) for v in vals):
        return False
    return True


def _is_exception_row(row: dict) -> bool:
    """Listed separately: parenthetical customer country and/or postcodes."""
    if not _is_data_country_row(row):
        return False
    cc = _str(row.get("CustomerCountry"))
    if "(" in cc or ")" in cc:
        return True
    if _nonblank(row.get("PostCode1")) or _nonblank(row.get("PostCode2")):
        return True
    return False


def _service_key_sort(k: str):
    m = re.search(r"(\d+)$", k, re.I)
    return int(m.group(1)) if m else 0


def _zone_sort_key(z: str):
    try:
        return (0, int(z))
    except ValueError:
        return (1, z)


def _format_exception_line(row: dict) -> str:
    parts = []
    if _nonblank(row.get("Country")):
        parts.append(f"Country={_str(row.get('Country'))}")
    if _nonblank(row.get("CustomerCountry")):
        parts.append(f"CustomerCountry={_str(row.get('CustomerCountry'))}")
    if _nonblank(row.get("PostCode1")):
        parts.append(f"PostCode1={_str(row.get('PostCode1'))}")
    if _nonblank(row.get("PostCode2")):
        parts.append(f"PostCode2={_str(row.get('PostCode2'))}")
    svc_keys = sorted(
        [k for k in row if _SERVICE_KEY.match(k)],
        key=_service_key_sort,
    )
    for k in svc_keys:
        v = _str(row.get(k))
        if v:
            parts.append(f"{k}={v}")
    if _nonblank(row.get("Country Code")):
        parts.append(f"Country Code={_str(row.get('Country Code'))}")
    return " | ".join(parts) if parts else repr(row)


def _split_standard_import_cz_country_region_line(line: str) -> list[str]:
    """
    When a Country Regions line lists ``Standard Import Zone <n> - …, CZ, …``,
    drop ``CZ`` from the list and add ``Standard Import Zone CZ<n> - CZ``.
    """
    s = line.strip()
    if not s:
        return [line]
    if re.match(r"(?i)^Service\d+:", s):
        return [line]
    m = re.match(
        r"^(?P<prefix>.+?\s+(?:Import|Export)\s+Zone\s+\d+)\s*-\s*(?P<codes>.+)$",
        s,
        re.I,
    )
    if not m:
        return [line]
    prefix = m.group("prefix").strip()
    if not re.search(r"(?i)^Standard\s+Import\s+Zone\s+\d+$", prefix):
        return [line]
    codes = [c.strip() for c in m.group("codes").split(",")]
    if "CZ" not in {c.upper() for c in codes if c}:
        return [line]
    zm = re.search(r"(?i)Zone\s+(\d+)\s*$", prefix)
    znum = zm.group(1) if zm else ""
    codes_wo_cz = [c for c in codes if c.upper() != "CZ"]
    return [
        f"{prefix} - {', '.join(codes_wo_cz)}",
        f"Standard Import Zone CZ{znum} - CZ",
    ]


def build_country_zoning_txt_lines_from_rows(rows: list) -> list:
    """
    Build TXT lines for CountryZoning from a list of row dicts (JSON or Excel-shaped).

    Returns a list of strings (no trailing newline on each line).
    """
    if not rows:
        return []

    header_end = _first_data_row_index(rows)
    header_rows = [r for r in rows[:header_end] if isinstance(r, dict)]
    split_idx = _find_export_import_split(rows)
    column_products = _find_column_product_names(rows, header_rows)

    # (ServiceN, zone_str) -> set of ISO/country codes (Code column)
    groups = defaultdict(set)
    case1_children = _case1_child_indices(rows)
    case2_clusters = _case2_clusters_map(rows)

    for idx, row in enumerate(rows):
        if not isinstance(row, dict):
            continue
        if not _is_data_country_row(row):
            continue
        if not _is_primary_zoning_row(row):
            continue
        if idx in case1_children:
            continue
        code = _str(row.get("Code") or row.get("Country Code")).strip().upper()
        if not code:
            continue
        ck2 = _case2_cluster_key_from_row(row)
        cluster_inds = case2_clusters.get(ck2) if ck2 is not None else None

        for k, v in row.items():
            if not _SERVICE_KEY.match(k):
                continue
            zs = _str(v)
            if not _is_zone_value(zs):
                continue
            if cluster_inds is not None:
                if not _case2_unanimous(cluster_inds, rows, k, zs):
                    continue
            groups[(k, zs)].add(code)

    lines = []
    service_keys = sorted({k for k, _ in groups}, key=_service_key_sort)

    for sk in service_keys:
        col_num = _service_column_number(sk)
        ei = _export_or_import(col_num, split_idx)
        product_raw = (column_products.get(sk) or "").strip()

        if product_raw:
            lines.append(f"{sk}: {product_raw.lower()} {ei.lower()}")
        else:
            lines.append(f"{sk}: {ei.lower()}")

        zone_keys = sorted({z for k, z in groups if k == sk}, key=_zone_sort_key)
        for z in zone_keys:
            codes = sorted(groups[(sk, z)])
            if product_raw:
                prefix = f"{_pretty_service_title(product_raw)} {ei} Zone {z}"
            else:
                prefix = f"{sk} {ei} Zone {z}"
            lines.append(f"{prefix} - {', '.join(codes)}")

        lines.append("")

    while lines and lines[-1] == "":
        lines.pop()

    out: list[str] = []
    for ln in lines:
        out.extend(_split_standard_import_cz_country_region_line(ln))
    return out


def _format_postal_exception_tabular(row: dict) -> str:
    """One line: Zone name - Country - Postal code - Excluded (non–Case-1 rows)."""
    z = _str(row.get("Country"))
    code = _str(row.get("Code") or row.get("Country Code")).strip().upper()
    pc1, pc2 = _str(row.get("PostCode1")), _str(row.get("PostCode2"))
    postal = ""
    if pc1 or pc2:
        postal = f"{pc1}-{pc2}".strip("-")
    cc = _str(row.get("CustomerCountry"))
    excluded = cc if ("(" in cc and ")" in cc) else ""
    return f"{z} - {code} - {postal} - {excluded}"


def build_postal_code_zones_txt_lines(rows: list) -> list:
    """
    Postal Code Zones file: Case 1–4 (Case 4 = ``PostCode1``/``PostCode2``),
    then other exception rows.

    Column layout per line: ``Zone name - Country - Postal code - Excluded``.
    """
    if not rows:
        return []

    header_end = _first_data_row_index(rows)
    header_rows = [r for r in rows[:header_end] if isinstance(r, dict)]
    split_idx = _find_export_import_split(rows)
    column_products = _find_column_product_names(rows, header_rows)
    case1_children = _case1_child_indices(rows)
    case2_clusters = _case2_clusters_map(rows)
    case2_all_indices: set[int] = set()
    for _ci in case2_clusters.values():
        case2_all_indices.update(_ci)

    lines: list[str] = [
        "Postal Code Zones",
        "Zone name - Country - Postal code - Excluded",
        "",
    ]

    by_cc_code: dict[tuple[str, str], list[tuple[int, dict]]] = defaultdict(list)
    for i, row in enumerate(rows):
        if not isinstance(row, dict) or not _is_data_country_row(row):
            continue
        cc = _norm_key(_str(row.get("CustomerCountry")))
        code = _str(row.get("Code") or row.get("Country Code")).strip().upper()
        if not code:
            continue
        by_cc_code[(cc, code)].append((i, row))

    processed_ids: set[int] = set()
    parent_indices: list[int] = []
    for i, row in enumerate(rows):
        if not isinstance(row, dict) or not _is_data_country_row(row):
            continue
        if not _country_has_excluding(_str(row.get("Country"))):
            continue
        if not _parse_excluded_regions_from_country(_str(row.get("Country"))):
            continue
        code = _str(row.get("Code") or row.get("Country Code")).strip().upper()
        if not code:
            continue
        parent_indices.append(i)

    case1_emitted = False
    for pi in parent_indices:
        row = rows[pi]
        excluded = _parse_excluded_regions_from_country(_str(row.get("Country")))
        if not excluded:
            continue
        code = _str(row.get("Code") or row.get("Country Code")).strip().upper()
        cc_key = _norm_key(_str(row.get("CustomerCountry")))

        children_pairs: list[tuple[int, dict]] = []
        seen_j: set[int] = set()
        for j, r2 in by_cc_code.get((cc_key, code), []):
            if j == pi or j in seen_j:
                continue
            c2 = _str(r2.get("Country"))
            if _country_has_excluding(c2):
                continue
            if _child_matches_excluded_token(c2, excluded):
                children_pairs.append((j, r2))
                seen_j.add(j)

        processed_ids.add(pi)
        for j, _ in children_pairs:
            processed_ids.add(j)

        if case1_emitted:
            lines.append("")
        case1_emitted = True

        svc_items = sorted(
            [(k, row[k]) for k in row if _SERVICE_KEY.match(k)],
            key=lambda kv: _service_key_sort(kv[0]),
        )
        for k, v in svc_items:
            zs = _str(v)
            if not _is_zone_value(zs):
                continue
            zt = _postal_zone_title(k, zs, code, split_idx, column_products, None)
            exc_txt = ", ".join(sorted(excluded))
            lines.append(
                f"{zt} - {code} - {_CASE1_MAINLAND_POSTAL_PLACEHOLDER} - {exc_txt}"
            )

        for j, ch in children_pairs:
            sub = _str(ch.get("Country")).strip()
            ch_items = sorted(
                [(k, ch[k]) for k in ch if _SERVICE_KEY.match(k)],
                key=lambda kv: _service_key_sort(kv[0]),
            )
            for k, v in ch_items:
                zs = _str(v)
                if not _is_zone_value(zs):
                    continue
                zt = _postal_zone_title(
                    k, zs, code, split_idx, column_products, sub
                )
                lines.append(f"{zt} - {code} - {sub} - ")

    case2_header_done = False
    for _ck in sorted(case2_clusters.keys(), key=lambda x: (x[0], x[1])):
        inds = case2_clusters[_ck]
        all_names = _case2_cluster_country_names(rows, inds)
        code = _str(rows[inds[0]].get("Code") or rows[inds[0]].get("Country Code"))
        code = code.strip().upper()
        if not code or not all_names:
            continue

        pair_to_countries: dict[tuple[str, str], set[str]] = defaultdict(set)
        for j in inds:
            r = rows[j]
            ctry = _str(r.get("Country")).strip()
            if not ctry:
                continue
            for k, v in r.items():
                if not _SERVICE_KEY.match(k):
                    continue
                zs = _str(v)
                if not _is_zone_value(zs):
                    continue
                pair_to_countries[(k, zs)].add(ctry)

        zone_pairs = sorted(
            pair_to_countries.keys(),
            key=lambda p: (_service_key_sort(p[0]), _zone_sort_key(p[1])),
        )
        for sk, zs in zone_pairs:
            if _case2_unanimous(inds, rows, sk, zs):
                continue
            subset = sorted(
                pair_to_countries[(sk, zs)],
                key=lambda x: x.lower(),
            )
            z_title = _case2_zone_title(
                sk,
                zs,
                code,
                split_idx,
                column_products,
                subset,
                all_names,
            )
            bare = _case2_use_bare_gb_title(subset, all_names)
            postal_col, ex_col = _case2_postal_and_excluded_columns(
                subset, all_names, bare
            )
            if not case2_header_done:
                if case1_emitted:
                    lines.append("")
                    lines.append("=" * 72)
                lines.append("Case 2 - same Code, different Country")
                lines.append("-" * 72)
                case2_header_done = True
            lines.append(f"{z_title} - {code} - {postal_col} - {ex_col}")

    case3_processed: set[int] = set()
    case3_groups = _case3_groups(rows)
    case3_header_done = False
    lead_sep_case3 = case1_emitted or case2_header_done

    def _sort_range_key(nr: str) -> int:
        try:
            return int(nr.split("-", 1)[0])
        except (ValueError, IndexError):
            return 0

    for _gk in sorted(case3_groups.keys(), key=lambda x: (x[0], x[1])):
        inds = case3_groups[_gk]
        r0 = rows[inds[0]]
        parsed0 = _parse_case3_customer_country(_str(r0.get("CustomerCountry")))
        base_display = (parsed0[0] if parsed0 else "").strip()
        code = _str(r0.get("Code") or r0.get("Country Code")).strip().upper()
        if not base_display or not code:
            continue

        numeric_ranges: list[str] = []
        other_indices: list[int] = []
        range_indices: list[tuple[int, str]] = []

        for j in inds:
            r = rows[j]
            pr = _parse_case3_customer_country(_str(r.get("CustomerCountry")))
            if not pr:
                continue
            _b, inner = pr
            if _case3_inner_is_other(inner):
                other_indices.append(j)
                continue
            nr = _case3_inner_numeric_range(inner)
            if nr:
                numeric_ranges.append(nr)
                range_indices.append((j, nr))

        numeric_ranges_uniq = sorted(set(numeric_ranges), key=_sort_range_key)
        range_indices_sorted = sorted(range_indices, key=lambda t: _sort_range_key(t[1]))

        if not range_indices_sorted and not other_indices:
            continue

        if not case3_header_done:
            if lead_sep_case3:
                lines.append("")
                lines.append("=" * 72)
            lines.append("Case 3 - postcodes in Customer country")
            lines.append("-" * 72)
            case3_header_done = True

        for j, nr in range_indices_sorted:
            zone_display = f"{base_display} ({nr})"
            lines.append(f"{zone_display} - {code} - {nr} - ")
            case3_processed.add(j)

        for oj in other_indices:
            zone_other = f"{base_display} (other)"
            exc = ", ".join(numeric_ranges_uniq)
            lines.append(
                f"{zone_other} - {code} - {_CASE1_MAINLAND_POSTAL_PLACEHOLDER} - {exc}"
            )
            case3_processed.add(oj)

    case4_processed: set[int] = set()
    case4_by_code = _case4_indices_by_code(rows)
    case4_header_done = False
    lead_case4 = case1_emitted or case2_header_done or case3_header_done

    def _range_sort_key_case4(s: str) -> int:
        try:
            return int(s.split("-", 1)[0])
        except (ValueError, IndexError):
            return 0

    for code in sorted(case4_by_code.keys()):
        inds = case4_by_code[code]
        by_sig: dict[tuple[tuple[str, str], ...], list[int]] = defaultdict(list)
        for i in inds:
            sig = _case4_service_signature(rows[i])
            by_sig[sig].append(i)
        block_sigs = sorted(by_sig.keys(), key=lambda s: min(by_sig[s]))

        accumulated_ranges: set[str] = set()

        for sig in block_sigs:
            block_i = sorted(by_sig[sig])
            for bi in block_i:
                case4_processed.add(bi)
            block_rows = [rows[j] for j in block_i]
            template = rows[block_i[0]]
            code_u = code.strip().upper()

            has_other = False
            range_parts: list[str] = []
            for r in block_rows:
                p1, p2 = r.get("PostCode1"), r.get("PostCode2")
                if _postcode_row_is_other_cells(_str(p1), _str(p2)):
                    has_other = True
                elif _postcode_cells_numeric_pair(_str(p1), _str(p2)):
                    range_parts.append(
                        _format_postcode_range_numeric(_str(p1), _str(p2))
                    )
                else:
                    has_other = True

            svc_items = sorted(
                [(k, template[k]) for k in template if _SERVICE_KEY.match(k)],
                key=lambda kv: _service_key_sort(kv[0]),
            )

            if has_other:
                exc = ", ".join(
                    sorted(accumulated_ranges, key=_range_sort_key_case4)
                )
                for k, v in svc_items:
                    zs = _str(v)
                    if not _is_zone_value(zs):
                        continue
                    zt = _case4_zone_title(
                        k, zs, code_u, split_idx, column_products
                    )
                    if not case4_header_done:
                        if lead_case4:
                            lines.append("")
                            lines.append("=" * 72)
                        lines.append("Case 4 - postcodes in PostCode columns")
                        lines.append("-" * 72)
                        case4_header_done = True
                        lead_case4 = False
                    lines.append(
                        f"{zt} - {code_u} - {_CASE1_MAINLAND_POSTAL_PLACEHOLDER} - {exc}"
                    )
            else:
                merged_postal = ", ".join(
                    sorted(set(range_parts), key=_range_sort_key_case4)
                )
                accumulated_ranges.update(range_parts)
                for k, v in svc_items:
                    zs = _str(v)
                    if not _is_zone_value(zs):
                        continue
                    zt = _case4_zone_title(
                        k, zs, code_u, split_idx, column_products
                    )
                    if not case4_header_done:
                        if lead_case4:
                            lines.append("")
                            lines.append("=" * 72)
                        lines.append("Case 4 - postcodes in PostCode columns")
                        lines.append("-" * 72)
                        case4_header_done = True
                        lead_case4 = False
                    lines.append(
                        f"{zt} - {code_u} - {merged_postal} - "
                    )

    lines.append("")
    lines.append("Other postal / exception rows")
    lines.append("-" * 72)

    other_any = False
    for i, row in enumerate(rows):
        if not isinstance(row, dict) or not _is_data_country_row(row):
            continue
        if i in processed_ids:
            continue
        if i in case2_all_indices:
            continue
        if i in case3_processed:
            continue
        if i in case4_processed:
            continue
        if not (
            _is_exception_row(row)
            or _country_has_excluding(_str(row.get("Country")))
            or i in case1_children
        ):
            continue
        lines.append(_format_postal_exception_tabular(row))
        other_any = True

    if not other_any:
        lines.append("(none)")

    while lines and lines[-1] == "":
        lines.pop()

    return lines


def _rows_from_excel_sheet(excel_path: Path, sheet_name: str) -> list:
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required. Install with: pip install openpyxl")

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet_name]
    raw = list(ws.iter_rows(values_only=True))
    wb.close()
    if not raw:
        return []
    headers = [str(h).strip() if h is not None else "" for h in raw[0]]
    rows = []
    for tup in raw[1:]:
        d = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            val = tup[i] if i < len(tup) else None
            if val is not None and val != "":
                d[h] = val
        rows.append(d)
    return rows


def _append_demand_surcharge_countries_lines(lines, demand_surcharge_countries, name_to_code):
    """
    Append DemandSurchargeCountries block: one line per zone
    ``DemandSurcharge_Origin_<Token>`` or ``DemandSurcharge_Destination_<Token>``,
    matching the document's Origin vs Destination country tables.
    """
    if not demand_surcharge_countries:
        return
    lines.append("")
    section = "origin"
    for row in demand_surcharge_countries:
        if not isinstance(row, dict):
            continue
        od = (
            row.get("Origin_Destination")
            or row.get("origin_destination")
            or row.get("Origin/Destination")
            or row.get("origin/destination")
        )
        if od and str(od).strip():
            t = str(od).strip().lower()
            if "origin" in t and "territor" in t:
                section = "origin"
                continue
            if "destination" in t and "territor" in t:
                section = "destination"
                continue
        zn = row.get("ZoneName") or row.get("zoneName")
        if zn is None or not str(zn).strip():
            continue
        zone_name = str(zn).strip()
        if section == "destination":
            prefix = demand_surcharge_destination_label(zone_name)
        else:
            prefix = demand_surcharge_origin_label(zone_name)
        raw = row.get("Countries") or row.get("countries") or ""
        if not isinstance(raw, str):
            raw = str(raw)
        codes = _gogreen_country_list_to_codes(raw, name_to_code) if raw.strip() else ""
        lines.append(f"{prefix}  {codes}")


def _append_gogreen_block_lines(lines, go_green_array, name_to_code):
    """Append GoGreen placeholder block definitions (codes) from GoGreenPlusCost JSON."""
    if not go_green_array:
        return
    gg_lines = build_gogreen_block_txt_lines(go_green_array, name_to_code)
    if not gg_lines:
        return
    lines.append("")
    lines.extend(gg_lines)


def create_country_region_txt(
    excel_path: str = "output/UPS_Rate_Cards.xlsx",
    sheet_name: str = "CountryZoning",
    output_path: str | None = None,
    extracted_json_path: str | None = None,
    postal_output_path: str | None = None,
) -> str:
    """
    Write Country Regions summary TXT and **Postal Code Zones** TXT.

    **CountryZoning body**

    Prefer ``CountryZoning`` rows from ``extracted_json_path`` when the file exists.
    Otherwise reads the same logical columns from the Excel sheet ``sheet_name``.

    Then appends DemandSurchargeCountries / GoGreenPlusCost blocks from JSON when present
    (Country Regions file only).

    ``postal_output_path`` defaults to ``Postal_Code_Zones.txt`` next to the Country Regions file.

    Returns the path to the written Country Regions ``.txt`` file.
    """
    excel_path = Path(excel_path)
    print(f"[*] TXT Debug: excel_path={excel_path}")
    print(f"[*] TXT Debug: sheet_name={sheet_name}")

    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    if output_path is None:
        output_dir = excel_path.parent
        output_path = output_dir / "CountryZoning_by_RateName.txt"
    else:
        output_path = Path(output_path)

    cz_rows = []
    jp = Path(extracted_json_path) if extracted_json_path else None
    if jp and jp.exists():
        try:
            with open(jp, encoding="utf-8") as jf:
                data = json.load(jf)
            cz_rows = data.get("CountryZoning") or []
            if not isinstance(cz_rows, list):
                cz_rows = []
            print(f"[*] TXT Debug: CountryZoning rows from JSON={len(cz_rows)}")
        except Exception as e:
            print(f"[WARN] TXT Debug: could not read JSON, falling back to Excel: {e}")
            cz_rows = []

    if not cz_rows:
        cz_rows = _rows_from_excel_sheet(excel_path, sheet_name)
        print(f"[*] TXT Debug: CountryZoning rows from Excel={len(cz_rows)}")

    lines = build_country_zoning_txt_lines_from_rows(cz_rows)
    print(f"[*] TXT Debug: CountryZoning body lines={len(lines)}")

    if extracted_json_path:
        jp = Path(extracted_json_path)
        if jp.exists():
            try:
                with open(jp, encoding="utf-8") as jf:
                    data = json.load(jf)
                name_to_code = _load_country_codes()
                dsc = data.get("DemandSurchargeCountries") or []
                if dsc:
                    before = len(lines)
                    _append_demand_surcharge_countries_lines(lines, dsc, name_to_code)
                    print(
                        f"[*] TXT Debug: appended DemandSurchargeCountries "
                        f"({len(lines) - before} lines incl. separator)"
                    )
                ggc = data.get("GoGreenPlusCost") or []
                if ggc:
                    before = len(lines)
                    _append_gogreen_block_lines(lines, ggc, name_to_code)
                    print(
                        f"[*] TXT Debug: appended GoGreenPlusCost blocks "
                        f"({len(lines) - before} lines incl. separator)"
                    )
            except Exception as e:
                print(f"[WARN] TXT Debug: could not append DemandSurchargeCountries / GoGreen: {e}")
        else:
            print(f"[WARN] TXT Debug: extracted_json_path not found: {jp}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"[OK] TXT Debug: wrote file {output_path}")

    if postal_output_path is None:
        postal_path = output_path.parent / "Postal_Code_Zones.txt"
    else:
        postal_path = Path(postal_output_path)
    postal_lines = build_postal_code_zones_txt_lines(cz_rows)
    postal_path.parent.mkdir(parents=True, exist_ok=True)
    postal_path.write_text("\n".join(postal_lines), encoding="utf-8")
    print(f"[OK] TXT Debug: wrote file {postal_path}")

    return str(output_path)


def main():
    script_dir = Path(__file__).resolve().parent
    excel_path = script_dir / "output" / "UPS_Rate_Cards.xlsx"
    output_path = script_dir / "output" / "CountryZoning_by_RateName.txt"

    print("Creating CountryZoning TXTs from UPS_Rate_Cards.xlsx...")
    extracted_json = script_dir / "processing" / "extracted_data.json"
    postal_path = script_dir / "output" / "Postal_Code_Zones.txt"
    out = create_country_region_txt(
        excel_path=str(excel_path),
        output_path=str(output_path),
        extracted_json_path=str(extracted_json) if extracted_json.exists() else None,
        postal_output_path=str(postal_path),
    )
    print(f"Saved: {out}")
    print(f"Saved: {postal_path}")


if __name__ == "__main__":
    main()
