"""Excel writers: MainCosts matrix sheet, flat tabs, Accessorial Costs."""

import re

from transform_main_costs import MAIN_COSTS_SHIPMENT_COLS


def _range_weight_to_leq_display(weight_str):
    """
    Convert a range weight header to "<= Y" format for display.
    Dot is decimal separator, comma is thousands separator.
    Examples: "30.1-70" -> "<= 70", "300.1 - 99,999" -> "<= 99999" (no dot on whole numbers).
    """
    if not weight_str:
        return weight_str
    s = str(weight_str).strip()
    parts = re.split(r'[-–\s]+', s)
    if len(parts) < 2:
        return weight_str
    # First part is start (e.g. 30.1), rest is end; "99" + "995" -> 99995
    end_str = ''.join(parts[1:])
    # Comma = thousands separator: remove commas. Dot = decimal (unchanged).
    # So "99,999" -> 99999, "99,999.0" -> 99999.0 -> display "<= 99999" (no dot for whole)
    try:
        end_val = float(end_str.replace(',', ''))
        if end_val != end_val:  # NaN
            return weight_str
        # Whole numbers: show without dot (99999 not 99999.0)
        if end_val == int(end_val):
            return f"<= {int(end_val)}"
        return f"<= {end_val}"
    except ValueError:
        return weight_str


# Reserved column layout if an Accessorial Costs tab is added back later (not used by the current pipeline).
ACCESSORIAL_COSTS_COLUMNS = [
    'Original Cost Name',          # the cost name as it appears in the rate card PDF
    'Cost Type',                   # standardised type name (filled by fuzzy matching)
    'Cost Price',                  # the numeric price value
    'Minimum',                     # extracted from "X with minimum of Y" in Cost Price
    'Currency',                    # e.g. EUR, USD
    'Rate by',                     # how the price is applied (e.g. per shipment, per kg)
    'Apply Over',                  # what the cost applies to (e.g. base freight)
    'Apply if',                    # condition under which the cost applies (left blank)
    'Additional info(Cost Code)',  # internal cost code from the rate card
    'Valid From',                  # start date of validity (taken from the rate card metadata)
    'Valid To',                    # end date of validity (not available; left blank)
    'Carrier',                     # carrier name
]


def write_matrix_sheet(workbook, sheet_name, matrix_rows, category_specs, metadata):
    """
    MainCosts: four header rows (merged transport-cost groups, weight measure, ``<=`` weights,
    currency/flat or adder labels), then data from row 5.
    """
    if not matrix_rows:
        print(f"[WARN] No matrix data for {sheet_name}, skipping")
        return

    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    print(f"[*] Creating {sheet_name} (Matrix) tab with {len(matrix_rows)} lanes...")
    ws = workbook.create_sheet(sheet_name)

    # Define the blue header style used for all three header rows
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Shipment / lane identity columns (see transform_main_costs.MAIN_COSTS_SHIPMENT_COLS)
    fixed_cols = list(MAIN_COSTS_SHIPMENT_COLS)
    num_fixed = len(fixed_cols)
    col = 1   # tracks the current column position as we build the header

    # --- Write fixed shipment column names in Row 1 ---
    for c, name in enumerate(fixed_cols, 1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    col = num_fixed + 1   # move the column pointer past the fixed columns

    # --- Build the cost category column groups (Rows 1–4) ---
    # category_specs: list of (cost_cat_name, blocks) where
    #   blocks = [(weight_unit, weights, row4_label), ...]
    # One category can have multiple blocks (e.g. main weights + adder columns).
    category_start_cols = []   # (start_col, end_col, cost_cat_name, weights) per block for data write

    for cost_cat_name, blocks in category_specs:
        cat_start_col = col   # first column of this whole category (for Row 1 merge)

        for weight_unit, weights, row4_label in blocks:
            is_adder = (row4_label != 'Flat')
            start_col = col

            if is_adder:
                # Adder block: no spacer column, no "Rate by: p/X unit" label (user requested).
                for w_idx, w in enumerate(weights):
                    c = start_col + w_idx
                    ws.cell(row=2, column=c, value='')
                    ws.cell(row=2, column=c).fill = header_fill
                    ws.cell(row=2, column=c).font = header_font
                    ws.cell(row=2, column=c).alignment = header_alignment
                    # Row 3: show range as "<= Y" (e.g. 30.1-70 -> <= 70, 300.1-99-995 -> <= 99995)
                    ws.cell(row=3, column=c, value=_range_weight_to_leq_display(w))
                    ws.cell(row=3, column=c).fill = header_fill
                    ws.cell(row=3, column=c).font = header_font
                    ws.cell(row=3, column=c).alignment = header_alignment
                    # Row 4: "p/X unit" (no "Currency")
                    ws.cell(row=4, column=c, value=row4_label)
                    ws.cell(row=4, column=c).fill = header_fill
                    ws.cell(row=4, column=c).font = header_font
                    ws.cell(row=4, column=c).alignment = header_alignment
                col = start_col + len(weights)
            else:
                # Normal block: spacer with "Rate by: Weight measure - KG", then weight columns
                _base_label = f"Weight measure - {weight_unit}" if weight_unit else "Weight measure"
                weight_measure_label = f"Rate by: {_base_label}"
                ws.cell(row=2, column=col, value=weight_measure_label)
                ws.cell(row=2, column=col).fill = header_fill
                ws.cell(row=2, column=col).font = header_font
                ws.cell(row=2, column=col).alignment = header_alignment
                col += 1
                for _ in weights:
                    ws.cell(row=2, column=col, value='')
                    ws.cell(row=2, column=col).fill = header_fill
                    col += 1
                ws.cell(row=3, column=start_col, value='')
                ws.cell(row=3, column=start_col).fill = header_fill
                col = start_col + 1
                for w in weights:
                    ws.cell(row=3, column=col, value=f"<= {w}")
                    ws.cell(row=3, column=col).fill = header_fill
                    ws.cell(row=3, column=col).font = header_font
                    ws.cell(row=3, column=col).alignment = header_alignment
                    col += 1
                end_col = col - 1
                # Row 4: spacer "Currency", then "Flat" under each weight
                ws.cell(row=4, column=start_col, value='Currency')
                ws.cell(row=4, column=start_col).fill = header_fill
                ws.cell(row=4, column=start_col).font = header_font
                ws.cell(row=4, column=start_col).alignment = header_alignment
                for w_idx in range(len(weights)):
                    c = start_col + 1 + w_idx
                    ws.cell(row=4, column=c, value=row4_label)
                    ws.cell(row=4, column=c).fill = header_fill
                    ws.cell(row=4, column=c).font = header_font
                    ws.cell(row=4, column=c).alignment = header_alignment

            end_col = col - 1
            category_start_cols.append((start_col, end_col, cost_cat_name, weights, not is_adder))

        cat_end_col = col - 1
        # Row 1: merge all columns for this category (all blocks) and write category name once
        if cat_start_col <= cat_end_col:
            ws.merge_cells(start_row=1, start_column=cat_start_col, end_row=1, end_column=cat_end_col)
            cell = ws.cell(row=1, column=cat_start_col, value=cost_cat_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

    total_cols = col - 1

    # Rows 2, 3 and 4 under the fixed shipment columns: empty cells with the blue header fill
    for r in (2, 3, 4):
        for c in range(1, num_fixed + 1):
            ws.cell(row=r, column=c, value='')
            ws.cell(row=r, column=c).fill = header_fill

    doc_currency = (metadata.get('document_currency') or '').strip()

    # --- Write the data rows starting at row 5 (shifted down by one for the new Currency row) ---
    for row_idx, row_data in enumerate(matrix_rows, 5):
        col = 1

        # Write fixed shipment columns (see MAIN_COSTS_SHIPMENT_COLS)
        for fc in fixed_cols:
            val = row_data.get(fc, '')
            cell = ws.cell(row=row_idx, column=col, value=val)
            if fc == 'Lane #':
                cell.alignment = Alignment(horizontal="center")   # numbers look better centred
            else:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            col += 1

        # Write the price columns for each block (adder blocks have no spacer column).
        for start_col, end_col, cost_cat_name, weights, has_spacer in category_start_cols:
            col = start_col
            if has_spacer:
                cell = ws.cell(row=row_idx, column=col, value=doc_currency)
                cell.alignment = Alignment(horizontal="center")
                col = start_col + 1
            for w in weights:
                val = row_data.get((cost_cat_name, w), '')
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.alignment = Alignment(horizontal="center")
                col += 1

    # --- Auto-size column widths ---
    # Sample the content of up to 53 rows (3 header rows + first 50 data rows)
    # to estimate a good column width.  Cap at 50 characters to avoid very wide columns.
    last_data_row = len(matrix_rows) + 3
    for c in range(1, total_cols + 1):
        col_letter = get_column_letter(c)
        max_len = 10   # minimum width
        for r in range(1, min(last_data_row + 1, 54)):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    # Freeze the first four rows so the header stays visible when scrolling down
    ws.freeze_panes = "A5"
    # Add a filter dropdown to row 4 (the Currency/Flat row) so users can filter
    ws.auto_filter.ref = f"A4:{get_column_letter(total_cols)}{last_data_row}"
    print(f"[OK] {sheet_name} (Matrix) tab created with {total_cols} columns")


def write_sheet(workbook, sheet_name, rows, metadata):
    """
    Write a standard flat-table Excel sheet (used for AddedRates, CountryZoning,
    AdditionalZoning, ZoningMatrix, AdditionalCostsPart1, AdditionalCostsPart2).

    This is the generic writer used for all tabs except MainCosts (which has its own
    special three-row header).  It produces a simple one-row header + data rows layout.

    COLUMN ORDERING:
    Columns are arranged in three groups, in this order:
      1. Priority columns  – always appear first in a fixed human-friendly sequence
                             (Client, Carrier, Validity Date, Country, Country Code, …)
      2. Weight columns    – columns whose name contains "KG", starts with "<=", or contains "-"
                             sorted numerically (0.5 KG before 1 KG before 2 KG)
      3. Zone columns      – "Zone 1", "Zone 2" … sorted numerically
         Other columns     – everything else, sorted alphabetically
    """
    if not rows:
        print(f"[WARN] No data for {sheet_name}, skipping")
        return

    print(f"[*] Creating {sheet_name} tab with {len(rows)} rows...")

    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    ws = workbook.create_sheet(sheet_name)

    # Collect every column name that appears in any row (some rows may have extra fields)
    all_columns = set()
    for row in rows:
        all_columns.update(row.keys())

    # -----------------------------------------------------------------------
    # Step 1: Place the priority columns first.
    # These are the most important / most commonly used columns and should
    # always appear on the left side of the sheet.
    # -----------------------------------------------------------------------
    priority_cols = [
        'Client', 'Carrier', 'Validity Date',   # identity columns (always first)
        'Section', 'Service Type', 'Cost Category', 'Weight Unit', 'Zone',
        'Page Stopper', 'Table Name', 'Weight From', 'Weight To',
        'Origin', 'Destination', 'Cost', 'Currency', 'Rate By', 'Service',
        'RateName', 'Country', 'Country Code', 'WeightFrom', 'WeightTo'
    ]

    columns = []
    for col in priority_cols:
        if col in all_columns:
            columns.append(col)
            all_columns.discard(col)   # remove from the remaining set so it doesn't appear twice

    # -----------------------------------------------------------------------
    # Step 2: From the remaining columns, separate weight columns from everything else.
    # Weight columns are identified by their name pattern:
    #   - Contains "KG"    e.g. "0.5 KG", "1 KG"
    #   - Starts with "<=" e.g. "<=0.5"
    #   - Contains "-"     e.g. "0-0.5"
    # -----------------------------------------------------------------------
    weight_cols = []
    other_cols = []

    for col in all_columns:
        if 'KG' in col or col.startswith('<=') or '-' in col:
            weight_cols.append(col)
        else:
            other_cols.append(col)

    # Sort weight columns numerically by the leading number
    # e.g. "0.5 KG", "1 KG", "2 KG" (not "0.5 KG", "2 KG", "1 KG")
    try:
        weight_cols_sorted = sorted(weight_cols, key=lambda x: float(x.split()[0]))
    except Exception:
        weight_cols_sorted = sorted(weight_cols)   # fallback: alphabetical

    # Sort "Zone N" columns numerically (Zone 1, Zone 2, Zone 10 …)
    # and sort all other columns alphabetically after them.
    def _other_col_sort_key(c):
        m = re.match(r'^Zone\s+(\d+)$', c, re.IGNORECASE)
        if m:
            return (0, int(m.group(1)))   # group 0: sort by zone number
        return (1, c)                      # group 1: sort alphabetically

    columns.extend(weight_cols_sorted)
    columns.extend(sorted(other_cols, key=_other_col_sort_key))

    # Define the blue header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Write the header row (row 1) with the column names
    for col_idx, column in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=column)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Write the data rows starting at row 2.
    # For each row, look up the value for each column and write it to the correct cell.
    # If a row doesn't have a value for a column, write an empty string.
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, column in enumerate(columns, 1):
            value = row_data.get(column, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Short numeric/code values are centred; longer text values wrap inside the cell
            if column in ['Weight', 'Weight Unit', 'Section', 'Zone', 'Currency', 'Rate'] or 'KG' in column:
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Auto-size column widths by looking at the content of the first 50 data rows.
    # The width is capped between 10 and 50 characters to avoid extremes.
    for col_idx, column in enumerate(columns, 1):
        col_letter = get_column_letter(col_idx)
        max_length = len(str(column))   # start with the header name length as the minimum
        for row_idx in range(2, min(len(rows) + 2, 52)):   # sample up to 50 data rows
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Freeze the header row so column names stay visible when scrolling down
    ws.freeze_panes = "A2"
    # Add filter dropdowns to every column so users can filter/sort the data
    ws.auto_filter.ref = ws.dimensions

    print(f"[OK] {sheet_name} tab created with {len(columns)} columns")


def write_accessorial_sheet(workbook, sheet_name, rows):
    """
    Write the Accessorial Costs tab to Excel.

    This is a simplified version of write_sheet() that uses the fixed column order
    defined in ACCESSORIAL_COSTS_COLUMNS instead of dynamically determining columns.
    The column order is fixed because the Accessorial Costs tab has a specific agreed layout.
    """
    if not rows:
        print(f"[WARN] No data for {sheet_name}, skipping")
        return

    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    print(f"[*] Creating {sheet_name} tab with {len(rows)} rows...")
    ws = workbook.create_sheet(sheet_name)
    columns = ACCESSORIAL_COSTS_COLUMNS   # use the fixed column list defined at the top of this file

    # Define the blue header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Write the header row (row 1) with the fixed column names
    for col_idx, column in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=column)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Write the data rows starting at row 2.
    # All cells use wrap_text so long cost names are readable without widening the column too much.
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, column in enumerate(columns, 1):
            value = row_data.get(column, '')   # empty string if this row has no value for this column
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Auto-size columns by sampling up to 100 data rows (more than write_sheet's 50,
    # because cost names can be long and we want to capture outliers)
    for col_idx, column in enumerate(columns, 1):
        col_letter = get_column_letter(col_idx)
        max_length = len(str(column))   # start with the header name length
        for row_idx in range(2, min(len(rows) + 2, 102)):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 10), 50)

    # Freeze the header row and add filter dropdowns
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    print(f"[OK] {sheet_name} tab created with {len(columns)} columns")
