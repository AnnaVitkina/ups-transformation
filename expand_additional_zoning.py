"""
Post-processing step: expand MainCosts lanes using AdditionalZoning data.

PROBLEM THIS SOLVES:
Some countries appear in the CountryZoning tab with a star suffix (e.g. "GROOT BRIT. (GB) *1",
"GROOT BRIT. (GB) *2"), meaning that country is split into sub-zones with different pricing.
The AdditionalZoning tab tells us which cities/regions belong to each sub-zone.

In MainCosts, these sub-zones appear as zone labels like "WW_EXP_IMP_ZONE_3".  A lane row
might look like:
    Origin: Netherlands  |  Destination: WW_EXP_IMP_ZONE_3  |  Service: DHL EXPRESS ...

But "GROOT BRIT. (GB) *1" maps to WW_EXP_IMP_ZONE_3 in CountryZoning, and its AdditionalInfo
says "LONDONDERRY (LDY), BELFAST (BFS)".

This module expands that single lane into additional rows — one per starred country entry —
filling plain **Origin Country** / **Destination Country** where the zone sat in
Origin Country Region / Destination Country Region.

OCCURRENCE MATCHING:
If "GROOT BRIT. (GB) *1" appears twice in AdditionalZoning, it also appears twice in
CountryZoning.  The Nth occurrence in AdditionalZoning matches the Nth occurrence in
CountryZoning (by order of appearance).

Public function:
    expand_main_costs_with_additional_zoning(xlsx_path, output_path=None)
        Reads the xlsx, performs the expansion, writes back (or to output_path).
        Returns the path of the written file.
"""

import re
from collections import defaultdict
from pathlib import Path


# ---------------------------------------------------------------------------
# Reading helpers
# ---------------------------------------------------------------------------

def _read_sheet_as_dicts(ws):
    """Read an openpyxl worksheet into a list of dicts using the first row as headers."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(h).strip() if h is not None else '' for h in rows[0]]
    data = []
    for row in rows[1:]:
        d = {}
        for h, v in zip(headers, row):
            d[h] = v
        data.append(d)
    return headers, data


def _read_main_costs_headers(ws):
    """
    Read the MainCosts sheet headers from rows 1-4 (the 4-row header structure).
    Returns (col_headers, data_start_row) where col_headers is a list of column
    header strings (one per column index, 0-based) and data_start_row is the
    1-based row index where data begins (row 5).
    """
    rows = list(ws.iter_rows(min_row=1, max_row=4, values_only=True))
    # Row 1 has the category names (merged across columns)
    # Row 3 has the weight breakpoints
    # For our purposes we just need the column count and the fixed column names
    row1 = rows[0] if rows else []
    col_count = len(row1)
    return col_count, 5   # data starts at row 5


# ---------------------------------------------------------------------------
# Normalization helpers (for matching zone labels and country keys across tabs)
# ---------------------------------------------------------------------------

def _normalize_zone_label(rate_name):
    """
    Normalize zone label so MainCosts (e.g. "WW_EXP_ZONE_2") matches CountryZoning
    (e.g. "WW_EXP_ZONE_Zone 2"). Replaces "_Zone N" or " Zone N" at end with "_N".
    """
    if not rate_name:
        return rate_name
    s = str(rate_name).strip()
    m = re.search(r'^(.+?)(?:_Zone| Zone)\s*(\d+|\w)\s*$', s, re.IGNORECASE)
    if m:
        return f"{m.group(1).rstrip('_')}_{m.group(2)}"
    return s


def _normalize_country_key(s):
    """
    Normalize country string for matching (e.g. "France (FR) *1" and "FRANCE (FR)*1"
    both become "FRANCE(FR)*1" by removing all spaces).
    """
    if not s:
        return ''
    return re.sub(r'\s+', '', str(s).strip().upper())


# ---------------------------------------------------------------------------
# Build lookups from CountryZoning and AdditionalZoning
# ---------------------------------------------------------------------------

def _build_additional_zoning_lookup(additional_zoning_rows):
    """
    Build a lookup: normalized_country_key -> [(country_display, additional_info), ...]
    preserving occurrence order. Uses normalized keys so "France (FR) *1" and
    "FRANCE (FR)*1" match.

    Rows with no Country but with AdditionalInfo are attached to the last seen Country.
    """
    lookup = defaultdict(list)
    last_country = None

    for row in additional_zoning_rows:
        country = (row.get('Country') or '').strip()
        info = (row.get('AdditionalInfo') or '').strip()

        if country and '*' in country:
            last_country = country
            norm_key = _normalize_country_key(country)
            if norm_key:
                lookup[norm_key].append((country, info))
        elif not country and info and last_country:
            norm_key = _normalize_country_key(last_country)
            if norm_key and lookup[norm_key]:
                last_display = lookup[norm_key][-1][0]
                lookup[norm_key].append((last_display, info))

    return lookup


def _build_country_zoning_lookup(country_zoning_rows):
    """
    Build a lookup: starred_country_key -> [(zone_label, zone_number), ...]
    preserving occurrence order (Nth entry matches Nth AdditionalZoning entry).

    e.g. "GROOT BRIT. (GB) *1" -> [("WW_EXP_IMP_ZONE_3", "3"), ("ECONOMY_EXP_IMP_ZONE_3", "3"), ...]
    """
    lookup = defaultdict(list)   # key -> list of (rate_name, zone) tuples

    for row in country_zoning_rows:
        country = (row.get('Country') or '').strip()
        rate_name = (row.get('RateName') or '').strip()
        zone = str(row.get('Zone') or '').strip()

        if country and '*' in country and rate_name:
            lookup[country].append((rate_name, zone))

    return lookup


# ---------------------------------------------------------------------------
# Core expansion logic
# ---------------------------------------------------------------------------

def _is_zone_label(value):
    """Return True if the value looks like a zone label (e.g. WW_EXP_IMP_ZONE_3)."""
    if not value:
        return False
    s = str(value).strip().upper()
    return bool(re.search(r'_ZONE_\w+$', s))


def _build_zones_with_only_starred_countries(country_zoning_rows):
    """
    From CountryZoning rows, build the set of zone labels where every country in that zone has *.
    If a zone contains ONLY starred countries, we will not add the base row in MainCosts
    (only the expanded rows from additional zoning).
    """
    zone_to_all_countries = defaultdict(list)
    last_rate_name = ''
    for row in country_zoning_rows:
        rate_name = (row.get('RateName') or '').strip()
        country = (row.get('Country') or '').strip()
        if rate_name:
            last_rate_name = rate_name
        if not country:
            continue
        zone_label = _normalize_zone_label(last_rate_name) if last_rate_name else ''
        if zone_label:
            zone_to_all_countries[zone_label].append(country)
    only_starred = set()
    for zl, countries in zone_to_all_countries.items():
        if countries and all('*' in str(c) for c in countries):
            only_starred.add(zl)
    return only_starred


def _build_zone_to_countries(cz_lookup, az_lookup):
    """
    Build a mapping: zone_label -> [(country_display, additional_info), ...]

    Uses normalized country keys so CountryZoning and AdditionalZoning match
    (e.g. "France (FR) *1" vs "FRANCE (FR)*1"). Uses _normalize_zone_label so
    "WW_EXP_ZONE_Zone 2" and "WW_EXP_ZONE_2" both resolve. Registers under both
    normalized and raw rate_name for robust lookup.
    """
    zone_to_countries = defaultdict(list)

    for country_key, cz_entries in cz_lookup.items():
        norm_country = _normalize_country_key(country_key)
        az_entries = az_lookup.get(norm_country, [])

        for idx, (rate_name, zone) in enumerate(cz_entries):
            if idx < len(az_entries):
                country_display, info = az_entries[idx]
            else:
                country_display, info = country_key, ''
            norm_zone = _normalize_zone_label(rate_name)
            zone_to_countries[norm_zone].append((country_display, info))
            if rate_name and rate_name != norm_zone:
                zone_to_countries[rate_name].append((country_display, info))

    return zone_to_countries


def expand_rows(main_costs_data_rows, zone_to_countries, carrier_country_code='', only_starred_zones=None):
    """
    For each MainCosts data row where Origin or Destination is a zone label,
    generate additional expanded rows — one per starred country entry for that zone.

    If only_starred_zones is set and the zone contains ONLY starred countries (from additional zoning),
    the base row is not added — only the expanded rows are added (additional zoning will cover it).
    Expanded rows set plain **Origin Country** / **Destination Country** as appropriate.

    Export lane: Origin = carrier, Destination = zone → when we expand destination,
    we fill Destination Country; the other side (Origin) is carrier → set Origin Country = carrier_country_code.
    Import lane: Origin = zone, Destination = carrier → when we expand origin,
    we fill Origin Country; the other side (Destination) is carrier → set Destination Country = carrier_country_code.

    Parameters:
        main_costs_data_rows   – list of dicts (one per data row, keyed by column index)
        zone_to_countries      – dict from _build_zone_to_countries()
        carrier_country_code   – ISO code for the carrier's country (e.g. 'DE'); used to fill
                                 the non-expanded side (Origin Country for export, Destination Country for import).
    Returns list of row dicts with the new columns populated where applicable.
    """
    result = []
    only_starred_zones = only_starred_zones or set()

    for row in main_costs_data_rows:
        origin = str(row.get('Origin Country Region') or row.get('Origin') or '').strip()
        destination = str(row.get('Destination Country Region') or row.get('Destination') or '').strip()

        origin_is_zone = _is_zone_label(origin)
        dest_is_zone = _is_zone_label(destination)

        # Determine which zone label to expand (if any)
        if origin_is_zone:
            zone_label = origin
            expand_field = 'origin'
        elif dest_is_zone:
            zone_label = destination
            expand_field = 'destination'
        else:
            zone_label = None
            expand_field = None

        # Look up using normalized label first (e.g. WW_EXP_ZONE_2), then raw
        entries = []
        if zone_label:
            entries = zone_to_countries.get(_normalize_zone_label(zone_label)) or zone_to_countries.get(zone_label, [])

        # If this zone contains ONLY starred countries, skip the base row (additional zoning rows will be added only)
        norm_zone = _normalize_zone_label(zone_label) if zone_label else ''
        skip_base_row = zone_label and norm_zone in only_starred_zones and entries

        if not skip_base_row:
            base_row = dict(row)
            base_row.setdefault('Origin Country', '')
            base_row.setdefault('Destination Country', '')
            result.append(base_row)

        if not entries:
            continue

        for country_display, _city_info in entries:
            new_row = dict(row)
            new_row.setdefault('Origin Country', '')
            new_row.setdefault('Destination Country', '')

            if expand_field == 'origin':
                # Import lane: we expanded origin (zone). Fill origin side; other side (destination) is carrier.
                new_row['Origin Country Region'] = ''
                new_row['Origin'] = ''
                new_row['Origin Country'] = country_display
                new_row['Destination Country'] = carrier_country_code or ''
            else:
                # Export lane: we expanded destination (zone). Fill destination side; other side (origin) is carrier.
                new_row['Destination Country Region'] = ''
                new_row['Destination'] = ''
                new_row['Destination Country'] = country_display
                new_row['Origin Country'] = carrier_country_code or ''

            result.append(new_row)

    return result


# ---------------------------------------------------------------------------
# Main entry point: read xlsx, expand, write back
# ---------------------------------------------------------------------------

def expand_main_costs_with_additional_zoning(xlsx_path, output_path=None):
    """
    Read the Excel file at xlsx_path, expand the MainCosts tab using AdditionalZoning
    and CountryZoning data, and write the result.

    If output_path is None, overwrites the input file.
    Returns the path of the written file.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    xlsx_path = Path(xlsx_path)
    output_path = Path(output_path) if output_path else xlsx_path

    print(f"[*] expand_additional_zoning: reading {xlsx_path.name}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # MainCosts is required; CountryZoning and AdditionalZoning are optional
    if 'MainCosts' not in wb.sheetnames:
        print(f"[WARN] expand_additional_zoning: sheet 'MainCosts' not found, skipping")
        return str(output_path)

    # --- Read CountryZoning and AdditionalZoning (optional) ---
    cz_rows = []
    az_rows = []
    if 'CountryZoning' in wb.sheetnames:
        _, cz_rows = _read_sheet_as_dicts(wb['CountryZoning'])
    if 'AdditionalZoning' in wb.sheetnames:
        _, az_rows = _read_sheet_as_dicts(wb['AdditionalZoning'])

    az_lookup = _build_additional_zoning_lookup(az_rows)
    cz_lookup = _build_country_zoning_lookup(cz_rows)
    zone_to_countries = _build_zone_to_countries(cz_lookup, az_lookup)
    only_starred_zones = _build_zones_with_only_starred_countries(cz_rows)

    if zone_to_countries:
        zone_keys = list(zone_to_countries.keys())
        print(f"[*] expand_additional_zoning: found {len(zone_keys)} zone labels to expand: {zone_keys[:20]}{'...' if len(zone_keys) > 20 else ''}")
        for zk in zone_keys[:8]:
            print(f"    '{zk}' -> {len(zone_to_countries[zk])} country/entries")
    else:
        print("[*] expand_additional_zoning: no starred-country zone mappings; applying carrier country code only")

    # --- Read MainCosts ---
    ws_mc = wb['MainCosts']
    all_rows = list(ws_mc.iter_rows(values_only=True))

    # Rows 1-4 are headers; data starts at row 5
    header_rows = all_rows[:4]
    data_rows_raw = all_rows[4:]

    from transform_main_costs import MAIN_COSTS_SHIPMENT_COLS

    # Fixed columns from build_matrix / write_matrix_sheet (includes plain country / city / custom zone)
    ORIG_FIXED = list(MAIN_COSTS_SHIPMENT_COLS)
    OUT_FIXED = ORIG_FIXED

    # Number of price columns = total cols minus fixed shipment cols
    col_count = len(all_rows[0]) if all_rows else 0
    price_col_count = max(0, col_count - len(ORIG_FIXED))   # number of price/category columns

    # Convert raw tuples to dicts: fixed cols by name, price cols by 0-based index offset
    def row_to_dict(raw_row):
        d = {}
        for i, v in enumerate(raw_row):
            if i < len(ORIG_FIXED):
                d[ORIG_FIXED[i]] = v
            else:
                d[i] = v   # price columns keyed by original index
        return d

    data_dicts = [row_to_dict(r) for r in data_rows_raw]

    # Debug: sample MainCosts rows with zone labels
    def _zone_side(d):
        o = str(d.get('Origin Country Region') or d.get('Origin') or '')
        t = str(d.get('Destination Country Region') or d.get('Destination') or '')
        return o, t

    zone_label_rows = []
    for d in data_dicts:
        o, t = _zone_side(d)
        if _is_zone_label(o) or _is_zone_label(t):
            zone_label_rows.append(d)
    if zone_label_rows:
        print(f"[*] expand_additional_zoning: MainCosts has {len(data_dicts)} rows; {len(zone_label_rows)} rows have a zone label in origin/destination region columns")
        origins = list({_zone_side(d)[0] for d in zone_label_rows[:12]})
        dests = list({_zone_side(d)[1] for d in zone_label_rows[:12]})
        print(f"    Sample Origin values: {origins[:8]}")
        print(f"    Sample Destination values: {dests[:8]}")

    # --- Resolve carrier country name -> ISO code from Metadata tab ---
    # For every row where Origin or Destination was filled by global_country()
    # (i.e. equals the carrier country name), copy the ISO code into
    # Origin Country / Destination Country.
    carrier_country_name = ''
    carrier_country_code = ''
    if 'Metadata' in wb.sheetnames:
        try:
            from transform_main_costs import global_country
            from transform_other_tabs import _load_country_codes, _country_to_code
            ws_meta = wb['Metadata']
            for r in ws_meta.iter_rows(values_only=True):
                if r and str(r[0] or '').strip().lower() == 'carrier':
                    carrier_val = str(r[1] or '').strip()
                    carrier_country_name = global_country({'carrier': carrier_val})
                    if carrier_country_name:
                        carrier_country_code = (
                            _country_to_code(carrier_country_name, _load_country_codes())
                            or carrier_country_name
                        )
                    print(f"[*] expand_additional_zoning: carrier '{carrier_country_name}' -> '{carrier_country_code}'")
                    break
        except Exception as e:
            print(f"[WARN] expand_additional_zoning: could not resolve carrier country code: {e}")

    if carrier_country_name and carrier_country_code:
        for d in data_dicts:
            ocr = str(d.get('Origin Country Region') or d.get('Origin') or '').strip()
            dcr = str(d.get('Destination Country Region') or d.get('Destination') or '').strip()
            if ocr == carrier_country_name:
                d['Origin Country'] = carrier_country_code
                d['Origin Country Region'] = ''
                d['Origin'] = ''
            if dcr == carrier_country_name:
                d['Destination Country'] = carrier_country_code
                d['Destination Country Region'] = ''
                d['Destination'] = ''

    expanded_dicts = expand_rows(data_dicts, zone_to_countries, carrier_country_code, only_starred_zones)

    added = len(expanded_dicts) - len(data_dicts)
    print(f"[*] expand_additional_zoning: {len(data_dicts)} original rows -> {len(expanded_dicts)} rows (+{added} expanded)")

    # Debug: sample expanded rows' Origin/Destination Country
    if added > 0:
        n0 = len(data_dicts)
        expanded_only = expanded_dicts[n0:] if len(expanded_dicts) > n0 else []
        for d in expanded_only[:6]:
            print(
                f"    Origin Country='{d.get('Origin Country', '')}'  "
                f"Destination Country='{d.get('Destination Country', '')}'"
            )

    # --- Replace starred country display strings with ISO codes ---
    # Values written by expand_rows into Origin Country / Destination Country come from
    # CountryZoning and look like "GROOT BRIT. (GB) *1" or "FRANKRIJK (FR) *2".
    # Strategy:
    #   1. Extract code from parentheses if present: "GROOT BRIT. (GB) *1" -> "GB"
    #   2. Otherwise strip the star suffix and look up the remaining name in
    #      dhl_country_codes.txt: "France *2" -> "France" -> "FR"
    try:
        from transform_other_tabs import _load_country_codes, _country_to_code
        _name_to_code = _load_country_codes()
    except Exception:
        _name_to_code = {}

    def _starred_to_code(value):
        if not value:
            return value
        s = str(value).strip()
        # Step 1: parenthetical code, e.g. "GROOT BRIT. (GB) *1" -> "GB"
        m = re.search(r'\(([A-Za-z]{2,3})\)', s)
        if m:
            return m.group(1).upper()
        # Step 2: strip star suffix (e.g. " *2") and look up name in country codes
        name = re.sub(r'\s*\*\d+\s*$', '', s).strip()
        if name and _name_to_code:
            code = _country_to_code(name, _name_to_code)
            if code:
                return code
        return s   # leave unchanged if nothing matched

    for d in expanded_dicts:
        for fc in OUT_FIXED:
            d.setdefault(fc, '')
        if d.get('Origin Country'):
            d['Origin Country'] = _starred_to_code(d['Origin Country'])
        if d.get('Destination Country'):
            d['Destination Country'] = _starred_to_code(d['Destination Country'])

    # Total output columns = new fixed cols + price cols
    new_col_count = len(OUT_FIXED) + price_col_count

    # --- Rebuild the MainCosts sheet ---
    sheet_idx = wb.sheetnames.index('MainCosts')
    del wb['MainCosts']
    ws_new = wb.create_sheet('MainCosts', sheet_idx)

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # --- Write header rows 1-4 ---
    # The first len(OUT_FIXED) columns get the new fixed column names in row 1,
    # blue fill in rows 2-4.
    # The remaining price columns are copied from the original header rows,
    # shifted right by the number of new columns added (len(OUT_FIXED) - len(ORIG_FIXED)).
    shift = len(OUT_FIXED) - len(ORIG_FIXED)   # how many extra fixed cols were inserted

    for row_idx, hrow in enumerate(header_rows, 1):
        # Write the new fixed column headers (row 1) or blue fill (rows 2-4)
        for col_idx, col_name in enumerate(OUT_FIXED, 1):
            if row_idx == 1:
                val = col_name
            else:
                val = ''
            cell = ws_new.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        # Copy the original price-column header cells, shifted right
        for orig_col_idx in range(len(ORIG_FIXED), col_count):
            orig_val = hrow[orig_col_idx] if orig_col_idx < len(hrow) else None
            new_col_idx = orig_col_idx + shift + 1   # 1-based
            cell = ws_new.cell(row=row_idx, column=new_col_idx, value=orig_val)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

    # --- Write data rows starting at row 5 ---
    def _cell_val(val):
        return str(val) if val is not None and val != '' else ''

    for row_idx, d in enumerate(expanded_dicts, 5):
        col = 1
        # New fixed column order
        for fc in OUT_FIXED:
            ws_new.cell(row=row_idx, column=col, value=_cell_val(d.get(fc, '')))
            col += 1
        # Price columns (keyed by original 0-based index, starting at len(ORIG_FIXED))
        for orig_idx in range(len(ORIG_FIXED), col_count):
            ws_new.cell(row=row_idx, column=col, value=_cell_val(d.get(orig_idx, '')))
            col += 1

    # Freeze panes and auto-filter
    from openpyxl.utils import get_column_letter
    ws_new.freeze_panes = 'A5'
    ws_new.auto_filter.ref = f"A4:{get_column_letter(new_col_count)}{4 + len(expanded_dicts)}"

    wb.save(output_path)
    print(f"[OK] expand_additional_zoning: saved to {output_path.name}")
    return str(output_path)


