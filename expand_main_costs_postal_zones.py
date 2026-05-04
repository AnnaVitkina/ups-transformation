"""
Post-process MainCosts: correlate UPS lane zone titles with Postal Code Zones TXT,
duplicate lanes with postal-specific shipment columns, and migrate CZ-style labels.

1. **Postal correlation:** For each lane whose Origin/Destination Country Region holds a
   semantic UPS zone title (e.g. ``Express Freight Import Zone 2``), find matching
   zone name lines in ``Postal_Code_Zones.txt`` (e.g. ``Import express freight DE 2``).
   Append copied lanes with costs unchanged; set **Postal Code Zone** on the same side
   to the full postal zone name and clear the Country Region cell that held the title.

2. **Standard Import CZ:** If **Postal Code Zone** contains
   ``Standard Import Zone <n> CZ<n>``, move **Standard Import Zone CZ<n>** to
   **Country Region** (same side) and clear postal.

Requires ``Postal_Code_Zones.txt`` next to the workbook (same directory). If missing,
only CZ migration runs when applicable.

Public:
    expand_main_costs_with_postal_zones(xlsx_path, postal_txt_path=None, output_path=None)
"""

from __future__ import annotations

import copy
import re
from pathlib import Path

from transform_main_costs import MAIN_COSTS_SHIPMENT_COLS

# MainCosts titles: "{Product} {Import|Export} Zone {suffix}"
_MC_ZONE_TITLE = re.compile(
    r"(?i)^(?P<prod>.+?)\s+(?P<dir>Import|Export)\s+Zone\s+(?P<suffix>\S+)\s*$"
)

_CZ_POSTAL_TO_REGION = re.compile(
    r"(?i)Standard\s+Import\s+Zone\s+(\d+)\s+(CZ\d+)\b"
)


def _parse_postal_code_zones_txt(text: str) -> list[dict]:
    """Parse ``Zone name - Country - Postal - Excluded`` lines into dicts."""
    out: list[dict] = []
    for raw in text.splitlines():
        line = raw.strip()
        if not line or line.startswith("Postal Code Zones") or line.startswith("Zone name"):
            continue
        if line.startswith("Case ") or line.startswith("Other postal"):
            continue
        if set(line) <= {"-", "="} or line.startswith("-") and len(line) < 6:
            continue
        parts = line.split(" - ", 3)
        if len(parts) < 4:
            continue
        zone_name, country, postal, excluded = (
            parts[0].strip(),
            parts[1].strip(),
            parts[2].strip(),
            parts[3].strip(),
        )
        if not zone_name or not country:
            continue
        out.append(
            {
                "zone_name": zone_name,
                "country": country,
                "postal": postal,
                "excluded": excluded,
            }
        )
    return out


def _maincosts_zone_match_key(title: str) -> tuple[str, str, str] | None:
    m = _MC_ZONE_TITLE.match((title or "").strip())
    if not m:
        return None
    prod = " ".join(m.group("prod").split()).lower()
    direction = m.group("dir").lower()
    suffix = m.group("suffix").strip()
    return (direction, prod, suffix)


def _postal_zone_match_key(zone_name: str) -> tuple[str, str, str, str] | None:
    """Parse ``Import express freight DE 2`` → direction, product, ISO2, zone token."""
    t = (zone_name or "").split()
    if len(t) < 4:
        return None
    d0 = t[0].lower()
    if d0 not in ("import", "export"):
        return None
    code = t[-2]
    ztok = t[-1]
    if len(code) != 2 or not code.isalpha():
        return None
    product = " ".join(t[1:-2]).lower()
    return (d0, product, code.upper(), ztok)


def _keys_match(
    mc: tuple[str, str, str], postal: tuple[str, str, str, str]
) -> bool:
    d1, p1, s1 = mc
    d2, p2, _code, z2 = postal
    if d1 != d2:
        return False
    if p1 != p2:
        return False
    if s1 != z2:
        return False
    return True


def _semantic_zone_title(s: str) -> bool:
    return bool(_MC_ZONE_TITLE.match((s or "").strip()))


def _find_postal_matches(
    zone_title: str, postal_entries: list[dict]
) -> list[dict]:
    mc = _maincosts_zone_match_key(zone_title)
    if not mc:
        return []
    hits = []
    seen = set()
    for e in postal_entries:
        zn = e.get("zone_name") or ""
        pk = _postal_zone_match_key(zn)
        if not pk:
            continue
        if _keys_match(mc, pk):
            key = zn
            if key not in seen:
                seen.add(key)
                hits.append(e)
    return hits


def _apply_cz_standard_import_migration(row: dict) -> None:
    """Move ``Standard Import Zone N CZN`` from Postal → Country Region as ``Standard Import Zone CZN``."""
    for side in ("Origin", "Destination"):
        pk = f"{side} Postal Code Zone"
        rk = f"{side} Country Region"
        v = (row.get(pk) or "").strip()
        if not v:
            continue
        m = _CZ_POSTAL_TO_REGION.search(v)
        if not m:
            continue
        znum, cztag = m.group(1), m.group(2)
        new_region = f"Standard Import Zone {cztag}"
        row[rk] = new_region
        row[pk] = ""


def _expand_row_with_postal_matches(
    row: dict, postal_entries: list[dict]
) -> list[dict]:
    """Return ``[original] + copies`` with postal columns filled; original unchanged."""
    extras: list[dict] = []
    ocr = (row.get("Origin Country Region") or "").strip()
    dcr = (row.get("Destination Country Region") or "").strip()

    if ocr and _semantic_zone_title(ocr):
        for e in _find_postal_matches(ocr, postal_entries):
            zn = (e.get("zone_name") or "").strip()
            if not zn:
                continue
            dup = copy.deepcopy(row)
            dup["Origin Country Region"] = ""
            dup["Origin Postal Code Zone"] = zn
            extras.append(dup)

    if dcr and _semantic_zone_title(dcr):
        for e in _find_postal_matches(dcr, postal_entries):
            zn = (e.get("zone_name") or "").strip()
            if not zn:
                continue
            dup = copy.deepcopy(row)
            dup["Destination Country Region"] = ""
            dup["Destination Postal Code Zone"] = zn
            extras.append(dup)

    if not extras:
        return [row]
    return [row] + extras


def expand_main_costs_dicts_with_postal_zones(
    data_dicts: list[dict],
    postal_txt_path: Path | None,
) -> list[dict]:
    """
    Apply CZ migration and postal-zone lane expansion to MainCosts row dicts.
    Expects keys aligned with MAIN_COSTS_SHIPMENT_COLS + numeric price keys.
    """
    postal_entries: list[dict] = []
    if postal_txt_path and postal_txt_path.is_file():
        text = postal_txt_path.read_text(encoding="utf-8")
        postal_entries = _parse_postal_code_zones_txt(text)
        print(
            f"[*] expand_postal_zones: loaded {len(postal_entries)} postal zone lines from {postal_txt_path.name}"
        )
    else:
        print(
            f"[*] expand_postal_zones: no postal file at {postal_txt_path!r}; skipping postal match expansion"
        )

    expanded: list[dict] = []
    for d in data_dicts:
        d = copy.deepcopy(d)
        _apply_cz_standard_import_migration(d)
        if postal_entries:
            parts = _expand_row_with_postal_matches(d, postal_entries)
            expanded.extend(parts)
        else:
            expanded.append(d)

    for lane, row in enumerate(expanded, 1):
        row["Lane #"] = lane

    return expanded


def expand_main_costs_with_postal_zones(
    xlsx_path: str | Path,
    postal_txt_path: str | Path | None = None,
    output_path: str | Path | None = None,
) -> str:
    """
    Read workbook, expand MainCosts rows, write back.
    Default postal path: ``<xlsx_dir>/Postal_Code_Zones.txt``.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    xlsx_path = Path(xlsx_path)
    output_path = Path(output_path) if output_path else xlsx_path
    if postal_txt_path is None:
        postal_txt_path = xlsx_path.parent / "Postal_Code_Zones.txt"
    else:
        postal_txt_path = Path(postal_txt_path)

    print(f"[*] expand_postal_zones: reading {xlsx_path.name}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "MainCosts" not in wb.sheetnames:
        print("[WARN] expand_postal_zones: sheet 'MainCosts' not found, skipping")
        return str(output_path)

    ws_mc = wb["MainCosts"]
    all_rows = list(ws_mc.iter_rows(values_only=True))
    if not all_rows:
        wb.save(output_path)
        return str(output_path)

    header_rows = all_rows[:4]
    data_rows_raw = all_rows[4:]
    col_count = len(all_rows[0])
    ORIG_FIXED = list(MAIN_COSTS_SHIPMENT_COLS)

    def row_to_dict(raw_row):
        d = {}
        for i, v in enumerate(raw_row):
            if i < len(ORIG_FIXED):
                d[ORIG_FIXED[i]] = v
            else:
                d[i] = v
        return d

    data_dicts = [row_to_dict(r) for r in data_rows_raw]
    expanded_dicts = expand_main_costs_dicts_with_postal_zones(
        data_dicts, postal_txt_path
    )

    added = len(expanded_dicts) - len(data_dicts)
    print(
        f"[*] expand_postal_zones: {len(data_dicts)} rows -> {len(expanded_dicts)} (+{added})"
    )

    price_col_count = max(0, col_count - len(ORIG_FIXED))
    OUT_FIXED = ORIG_FIXED
    new_col_count = len(OUT_FIXED) + price_col_count
    shift = 0

    sheet_idx = wb.sheetnames.index("MainCosts")
    del wb["MainCosts"]
    ws_new = wb.create_sheet("MainCosts", sheet_idx)

    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, hrow in enumerate(header_rows, 1):
        for col_idx, col_name in enumerate(OUT_FIXED, 1):
            val = col_name if row_idx == 1 else ""
            cell = ws_new.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
        for orig_col_idx in range(len(ORIG_FIXED), col_count):
            orig_val = hrow[orig_col_idx] if orig_col_idx < len(hrow) else None
            new_col_idx = orig_col_idx + shift + 1
            cell = ws_new.cell(row=row_idx, column=new_col_idx, value=orig_val)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

    def _cell_val(val):
        return val if val is not None else ""

    for row_idx, d in enumerate(expanded_dicts, 5):
        col = 1
        for fc in OUT_FIXED:
            ws_new.cell(row=row_idx, column=col, value=_cell_val(d.get(fc)))
            col += 1
        for orig_idx in range(len(ORIG_FIXED), col_count):
            ws_new.cell(row=row_idx, column=col, value=_cell_val(d.get(orig_idx)))
            col += 1

    ws_new.freeze_panes = "A5"
    ws_new.auto_filter.ref = (
        f"A4:{get_column_letter(new_col_count)}{4 + len(expanded_dicts)}"
    )

    wb.save(output_path)
    print(f"[OK] expand_postal_zones: saved to {output_path.name}")
    return str(output_path)
