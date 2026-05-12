"""
Single pipeline entrypoint for the full DHL rate-card flow.

HOW THIS FILE FITS INTO THE BIGGER PICTURE
-------------------------------------------
This is the "master controller" script.  Running it kicks off the entire process
from start to finish in one go:

  1. You pick a Rate Card workbook (.xlsb / .xlsx) or a JSON file produced by conversion-to-json.py.
  2. The script loads structured pricing data via conversion-to-json.py (workbook_to_payload).
  3. It builds a formatted Excel workbook (``transformation_to_excel``, imported here as ``create_table``).
  4. It creates a CountryZoning summary TXT file (via country_region_txt_creation.py).
  5. It moves the processed input file to an archive folder so it doesn't get processed twice.

Designed to run both locally (Windows) and in Google Colab (cloud notebook).
"""

# --- Standard library imports ---
import argparse    # used to read command-line arguments (--input-file, --output-dir, etc.)
import contextlib  # used to temporarily redirect print output when running in quiet mode
import io          # used to capture print output as a string buffer (for quiet mode)
import json        # used to save the extracted JSON back to disk after service-type filling
import os          # used to read environment variables and check file sizes
import shutil      # used to copy and move files (archive, staging reference files)
import sys         # used to add the project root to Python's module search path
from pathlib import Path   # cross-platform file path handling


# ---------------------------------------------------------------------------
# HARDCODED DEFAULT PATHS
# ---------------------------------------------------------------------------
# These paths are used as fallbacks when no path is given on the command line
# and no environment variable is set.
#
# PRIORITY ORDER: command-line argument > environment variable > hardcoded default
#
# Two sets of paths are defined:
#   - Drive paths: used when running in Google Colab with Google Drive mounted
#   - Local paths: used when running on a local Windows machine

# --- Google Drive paths (Colab legacy defaults; none of these are required) ---
HARDCODED_INPUT_FOLDER = "/content/drive/Shareddrives/FA Ops Europe: Rate Maintenance Team /Documents/AI Adoption RMT/RMT UPS/input json"
HARDCODED_ARCHIVE_FOLDER = "/content/drive/Shareddrives/FA Ops Europe: Rate Maintenance Team /Documents/AI Adoption RMT/RMT UPS/archive"
# Optional: only used when CLIENTS_FILE / --clients-file is unset and the file exists on Drive.
HARDCODED_CLIENTS_FILE = "/content/drive/Shareddrives/FA Ops Europe: Rate Maintenance Team /Documents/AI Adoption RMT/RMT UPS/addition/clients.txt"
# Optional: only used when COUNTRY_CODES_FILE / --country-codes-file is unset and the file exists.
HARDCODED_COUNTRY_CODES_FILE = "/content/drive/Shareddrives/FA Ops Europe: Rate Maintenance Team /Documents/AI Adoption RMT/RMT UPS/addition/dhl_country_codes.txt"
HARDCODED_OUTPUT_DIR = "/content/drive/Shareddrives/FA Ops Europe: Rate Maintenance Team /Documents/AI Adoption RMT/RMT UPS/output"

# --- Local Windows fallbacks (used when Drive is not mounted). Files are optional. ---
LOCAL_INPUT_FOLDER = r"C:\Users\avitkin\.cursor\projects_folders\RMT\ups-transformation\input"
LOCAL_ARCHIVE_FOLDER = r"C:\Users\avitkin\.cursor\projects_folders\RMT\ups-transformation\archive"
LOCAL_CLIENTS_FILE = r"C:\Users\avitkin\.cursor\projects_folders\RMT\ups-transformation\addition\clients.txt"
LOCAL_COUNTRY_CODES_FILE = r"C:\Users\avitkin\.cursor\projects_folders\RMT\ups-transformation\addition\dhl_country_codes.txt"
LOCAL_OUTPUT_DIR = r"C:\Users\avitkin\.cursor\projects_folders\RMT\ups-transformation\output"


def _drive_available():
    """
    Check whether Google Drive is mounted and the Drive input folder exists.
    Returns True when running in Colab with Drive mounted; False on a local machine.
    This is used to decide which set of paths (Drive vs local) to use.
    """
    p = Path(HARDCODED_INPUT_FOLDER)
    return p.exists() and p.is_dir()


def _use_drive_or_local(path_str, local_fallback, is_dir=False):
    """
    Choose between a Drive path and a local fallback path.

    If path_str points to something that actually exists on disk, use it.
    Otherwise fall back to local_fallback (the Windows path).
    This lets the same code work on both Colab and local machines without changes.

    is_dir=True means we check that the path is a folder (not just a file).
    """
    if path_str:
        p = Path(path_str)
        if is_dir:
            if p.exists() and p.is_dir():
                return path_str   # Drive folder exists; use it
        else:
            if p.exists():
                return path_str   # Drive file exists; use it
    if local_fallback:
        print(f"[*] Using local path (Drive not available): {local_fallback}")
    return local_fallback or path_str   # fall back to local path


def _detect_project_root():
    """
    Find the root folder of the project (the folder that contains conversion-to-json.py and create_table).

    This is needed because the script can be run from different working directories
    (e.g. directly, via Colab exec(), or from a subfolder).  We try several candidate
    locations and return the first one that looks like the project root.

    Falls back to the current working directory if nothing else matches.
    """
    candidates = []

    # 1. Check if the REPO_ROOT environment variable is set (explicit override)
    env_root = os.environ.get("REPO_ROOT")
    if env_root:
        candidates.append(Path(env_root))

    # 2. Use the folder containing this script file (works for normal runs)
    if "__file__" in globals():
        candidates.append(Path(__file__).resolve().parent)

    # 3. Use the current working directory
    candidates.append(Path.cwd())

    # 4. Try known Colab paths where the repo might be cloned
    candidates.append(Path("/content/transformation-rate"))
    candidates.append(Path("/content/transformation-rates"))

    # 5. Scan all subfolders of /content/ (Colab-friendly: repo could be in any subfolder)
    content_root = Path("/content")
    if content_root.exists():
        for child in content_root.iterdir():
            if child.is_dir():
                candidates.append(child)

    # Prefer repo layout: create_table shim + Rate Card converter (conversion-to-json.py).
    for c in candidates:
        has_ct = (c / "create_table.py").exists()
        has_conv = (c / "conversion-to-json.py").exists()
        if has_ct and has_conv:
            return c.resolve()
    for c in candidates:
        if (c / "create_table.py").exists() and (c / "main.py").exists():
            return c.resolve()

    return Path.cwd().resolve()   # nothing matched; use current directory as last resort


# Detect the project root once at import time and add it to Python's module search path.
# This ensures that "import create_table" works regardless of where the script is launched from.
PROJECT_ROOT = _detect_project_root()
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

# Import the other modules in this project
import importlib.util

import transformation_to_excel as create_table   # builds the Excel workbook from extracted JSON
import fill_service_types    # fills in missing service_type values in the extracted data
from country_region_txt_creation import create_country_region_txt   # creates the CountryZoning TXT file

# Client list helper only (no extraction through main.py).
from main import read_client_list

PIPELINE_INPUT_DIR = PROJECT_ROOT / "input"
_CONVERSION_JSON_MODULE = None


def _conversion_json_module():
    """Load conversion-to-json.py (hyphenated filename; not a regular package name)."""
    global _CONVERSION_JSON_MODULE
    if _CONVERSION_JSON_MODULE is not None:
        return _CONVERSION_JSON_MODULE
    path = PROJECT_ROOT / "conversion-to-json.py"
    if not path.is_file():
        raise FileNotFoundError(f"Expected Rate Card converter at: {path}")
    spec = importlib.util.spec_from_file_location("conversion_json", path)
    if spec is None or spec.loader is None:
        raise ImportError(f"Could not load spec for {path}")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    _CONVERSION_JSON_MODULE = mod
    return mod


def _detect_client_from_input_path(client_list: list[str], filepath: str) -> str:
    """Match longest client name against the file path / name (same idea as main.detect_client filename step)."""
    if not client_list:
        return "Unknown"
    from pathlib import Path as _Path

    stem = _Path(filepath).name
    for ext in (".json", ".pdf", ".xlsx", ".xls", ".xlsb", ".csv"):
        if stem.lower().endswith(ext):
            stem = stem[: -len(ext)]
    stem_lower = stem.lower()
    for name in sorted(client_list, key=len, reverse=True):
        if name.lower() in stem_lower:
            return name
    import re as _re

    _STOP = {"rc", "rate", "rates", "ratecard", "card", "dhl", "express", "pdf", "toolbox", "gri", "non", "legacy"}
    words = _re.split(r"[\s_\-]+", stem)
    client_words = []
    for w in words:
        if w.lower() in _STOP or not w:
            break
        if w.isdigit():
            break
        client_words.append(w)
    if client_words:
        return " ".join(client_words).strip()
    return client_list[0]


def load_extracted_payload_from_input(input_file: str, client_list: list[str]) -> tuple[dict, str]:
    """
    Build the same ``extracted_data``-shape dict that downstream (fill_service_types, Excel) expects.

    - ``.xlsb`` / ``.xlsx``: ``conversion-to-json.workbook_to_payload``.
    - ``.json``: must already be extracted_data (top-level ``MainCosts``); not raw Azure DI JSON.
    """
    p = Path(input_file)
    if not p.is_file():
        raise FileNotFoundError(f"Input not found: {input_file}")
    suff = p.suffix.lower()
    client_guess = _detect_client_from_input_path(client_list, str(p))

    if suff in (".xlsb", ".xlsx"):
        conv = _conversion_json_module()
        # Let conversion-to-json derive metadata.client from the sheet (column A / Client:);
        # filename-based client_guess is only a fallback on the returned tuple.
        processed = conv.workbook_to_payload(
            p,
            client="Unknown",
            carrier=None,
            validity_date=None,
            document_currency="",
        )
        meta_client = (processed.get("metadata") or {}).get("client") or client_guess
        return processed, meta_client or "Unknown"

    if suff == ".json":
        with open(p, encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, dict):
            raise ValueError("JSON root must be an object")
        if "analyzeResult" in data:
            raise ValueError(
                "This pipeline expects Rate Card Excel or JSON from conversion-to-json.py, "
                "not raw Azure Document Intelligence JSON. Use: python conversion-to-json.py "
                "--input path/to/card.xlsb --output processing/from_rate_card_excel.json"
            )
        if "MainCosts" not in data:
            raise ValueError(
                "JSON must contain top-level 'MainCosts' (run conversion-to-json.py on the workbook first)."
            )
        meta = data.setdefault("metadata", {})
        if not meta.get("client"):
            meta["client"] = client_guess or "Unknown"
        if not meta.get("FileName"):
            meta["FileName"] = p.name
        return data, meta.get("client") or client_guess or "Unknown"

    raise ValueError(f"Unsupported input type {suff!r}. Use .xlsb, .xlsx, or extracted .json.")


def _save_extracted_json(data: dict, output_path: str) -> None:
    """Write extracted_data-shaped dict to JSON (same options as main.save_output)."""
    print(f"[*] Saving extracted data to: {output_path}")
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    kb = os.path.getsize(output_path) / 1024
    print(f"[OK] Saved ({kb:.2f} KB)")


def choose_pipeline_input_interactive() -> str:
    """Pick a Rate Card workbook or extracted JSON from ``input/`` (or default workbook from converter)."""
    inp = PIPELINE_INPUT_DIR
    files: list[Path] = []
    if inp.is_dir():
        for pat in ("*.json", "*.xlsb", "*.xlsx"):
            files.extend(inp.glob(pat))
        files = sorted({f.resolve(): f for f in files}.values(), key=lambda x: x.name.lower())
    if not files:
        conv = _conversion_json_module()
        default_wb = conv.default_input_workbook(PROJECT_ROOT)
        if default_wb is not None and default_wb.is_file():
            print(f"[WARN] No candidates in {inp}; using default workbook: {default_wb}")
            return str(default_wb)
        raise FileNotFoundError(
            f"No .json / .xlsb / .xlsx in {inp} and no default workbook found. "
            "Add a Rate Card file or set --input-file."
        )
    print("Select input file to process:")
    print()
    for i, path in enumerate(files, 1):
        size_mb = path.stat().st_size / (1024 * 1024)
        print(f"  {i}. {path.name}  ({size_mb:.2f} MB)")
    print()
    while True:
        choice = input(f"Enter number (1-{len(files)}): ").strip()
        try:
            n = int(choice)
            if 1 <= n <= len(files):
                return str(files[n - 1])
        except ValueError:
            pass
        print("Invalid choice. Enter a number from the list.")


def parse_args():
    """
    Read command-line arguments when the script is run from a terminal.

    All arguments are optional.  If none are provided, the script will either
    use hardcoded defaults or ask the user to choose a file interactively.

    Supported arguments:
      --input-file        path to Rate Card .xlsb / .xlsx or extracted JSON (from conversion-to-json.py)
      --input-folder      folder to list .json / .xlsb / .xlsx from (interactive pick)
      --archive-folder    where to move the processed input file after completion
      --clients-file      path to the clients.txt file (one client name per line)
      --country-codes-file path to the country code lookup file
      --output-dir        where to save the Excel and TXT output files
      --verbose           show full debug output from all sub-steps
    """
    parser = argparse.ArgumentParser(
        description="Run end-to-end Rate Card pipeline (conversion-to-json → Excel → TXT)."
    )
    parser.add_argument(
        "--input-file",
        default=None,
        help="Rate Card workbook (.xlsb/.xlsx) or extracted JSON path (can be on Google Drive).",
    )
    parser.add_argument(
        "--input-folder",
        default=None,
        help="Folder containing .json / .xlsb / .xlsx. Script lists them and asks you to choose one.",
    )
    parser.add_argument(
        "--archive-folder",
        default=None,
        help="Archive folder for the processed input file (.json / .xlsb / .xlsx). Default: <input-folder>/archive",
    )
    parser.add_argument(
        "--clients-file",
        default=None,
        help="Optional: clients file (one name per line). If omitted, uses addition/clients.txt when present.",
    )
    parser.add_argument(
        "--country-codes-file",
        default=None,
        help="Optional: country codes file (Country<TAB>Code). Not required for UPS pipeline.",
    )
    parser.add_argument(
        "--output-dir",
        default=None,
        help="Directory to write outputs (xlsx, txt). Extracted JSON is saved to processing/.",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Show full debug output from underlying modules.",
    )
    # parse_known_args is used instead of parse_args so that extra flags injected by
    # Colab (e.g. --f=...) don't cause the script to crash
    args, _unknown = parser.parse_known_args()
    return args


def _list_pipeline_input_files(folder_path):
    """Return sorted .json / .xlsb / .xlsx (Rate Card or extracted JSON)."""
    folder = Path(folder_path)
    if not folder.exists() or not folder.is_dir():
        raise FileNotFoundError(f"Input folder not found: {folder}")
    files: list[Path] = []
    for pat in ("*.json", "*.xlsb", "*.xlsx"):
        files.extend(folder.glob(pat))
    files = sorted({f.resolve(): f for f in files}.values(), key=lambda p: p.name.lower())
    if not files:
        raise FileNotFoundError(f"No .json / .xlsb / .xlsx in: {folder}")
    return files


def _choose_input_from_folder(folder_path):
    """Show numbered list of candidate inputs; return chosen Path."""
    files = _list_pipeline_input_files(folder_path)
    print("Select input file:")
    print()
    for i, p in enumerate(files, 1):
        size_mb = p.stat().st_size / (1024 * 1024)   # convert bytes to megabytes
        print(f"  {i}. {p.name}  ({size_mb:.2f} MB)")
    print()
    while True:
        choice = input(f"Enter number (1-{len(files)}): ").strip()
        try:
            n = int(choice)
            if 1 <= n <= len(files):
                return files[n - 1]   # return the chosen file path
        except ValueError:
            pass   # user typed something that isn't a number; ask again
        print("Invalid choice. Enter a number from the list.")


def resolve_input_file(input_arg, input_folder_arg=None):
    """
    Resolve input: Rate Card ``.xlsb`` / ``.xlsx`` or extracted-data ``.json``
    (from conversion-to-json.py). Bare filename is resolved under ``<project>/input/``.
    """
    if input_arg is None:
        env_input_folder = os.environ.get("INPUT_FOLDER")
        folder = input_folder_arg or env_input_folder or HARDCODED_INPUT_FOLDER
        if folder:
            folder = _use_drive_or_local(folder, LOCAL_INPUT_FOLDER, is_dir=True)
            selected = _choose_input_from_folder(folder)
            return str(selected), str(Path(folder))
        env_input = os.environ.get("INPUT_FILE")
        if env_input:
            return env_input, None
        return choose_pipeline_input_interactive(), None

    p = Path(input_arg)
    if not p.is_absolute() and len(p.parts) == 1:
        return str(PIPELINE_INPUT_DIR / p), None
    return str(p), None


def _archive_processed_input(input_file, input_folder=None, archive_folder=None):
    """
    Move the processed input file (workbook or JSON) to an archive folder so it will not be
    accidentally processed again in the future.

    Archive folder resolution order:
      1. Use archive_folder if provided.
      2. Check the ARCHIVE_FOLDER environment variable.
      3. Use the hardcoded default (Drive or local).
      4. Use <input_folder>/archive if input_folder is known.
      5. If none of the above, do nothing (return None).

    If a file with the same name already exists in the archive, a number suffix
    is added (e.g. myfile_1.json, myfile_2.json) to avoid overwriting.

    Returns the path where the file was archived, or None if archiving was skipped.
    """
    if archive_folder is None:
        archive_folder = os.environ.get("ARCHIVE_FOLDER")
    if archive_folder is None:
        archive_folder = HARDCODED_ARCHIVE_FOLDER
    if archive_folder is None and input_folder:
        archive_folder = str(Path(input_folder) / "archive")
    if not archive_folder:
        return None   # no archive location configured; skip archiving

    src = Path(input_file)
    if not src.exists():
        return None   # file already gone; nothing to archive

    archive_dir = Path(archive_folder)
    archive_dir.mkdir(parents=True, exist_ok=True)   # create archive folder if needed
    dst = archive_dir / src.name

    if dst.exists():
        # A file with this name already exists in the archive.
        # Add a numeric suffix to avoid overwriting it: myfile_1.json, myfile_2.json, ...
        stem = src.stem
        suffix = src.suffix
        i = 1
        while True:
            candidate = archive_dir / f"{stem}_{i}{suffix}"
            if not candidate.exists():
                dst = candidate
                break
            i += 1

    shutil.move(str(src), str(dst))   # move (not copy) the file to the archive
    return str(dst)


def _prepare_reference_files(country_codes_file):
    """
    Optionally stage ``dhl_country_codes.txt`` under ``input/`` for downstream lookups.

    This pipeline does **not** require a country-codes file (UPS flow). When omitted or
    missing, extraction still runs; ``transform_other_tabs`` falls back when no file is found.

    If the provided file is already under ``input/`` or ``addition/``, no copy is made.
    Otherwise the file is copied to ``input/dhl_country_codes.txt``.
    """
    if not country_codes_file:
        print(
            "[*] Country codes file not set; skipping (optional — set "
            "COUNTRY_CODES_FILE or --country-codes-file if needed)."
        )
        return
    src = Path(country_codes_file)
    if not src.exists():
        print(
            f"[WARN] Country codes file not found ({src}); continuing without it."
        )
        return
    in_input = (PROJECT_ROOT / "input" / "dhl_country_codes.txt").resolve() == src.resolve()
    in_addition = (PROJECT_ROOT / "addition" / "dhl_country_codes.txt").resolve() == src.resolve()
    if in_input or in_addition:
        print(f"[OK] Country codes used in place: {src}")
        return
    dst_dir = PROJECT_ROOT / "input"
    dst_dir.mkdir(parents=True, exist_ok=True)
    dst = dst_dir / "dhl_country_codes.txt"
    shutil.copy2(src, dst)
    print(f"[OK] Country codes staged: {dst}")


def run_pipeline(
    input_file,
    clients_file=None,
    country_codes_file=None,
    output_dir=None,
    input_folder=None,
    archive_folder=None,
    verbose=False,
):
    """
    Execute the full end-to-end pipeline for one Rate Card workbook or extracted JSON.

    Sequential steps:
      Step 1: Read the client list (metadata hint).
      Step 2: Load data via conversion-to-json.py (workbook_to_payload or pre-built JSON).
      Step 3: Fill any missing service_type values.
      Step 4: Build the Excel workbook from the extracted data.
      Step 5: Build the CountryZoning TXT summary file from the Excel workbook.

    After completion, the input file is moved to the archive folder.

    Parameters:
      input_file          – path to .xlsb / .xlsx Rate Card or extracted-data .json
      clients_file        – path to clients.txt (one client name per line)
      country_codes_file  – optional path to the country code lookup file
      output_dir          – folder where Excel and TXT outputs will be saved
      input_folder        – folder the input file came from (used for archiving)
      archive_folder      – where to move the input file after processing
      verbose             – if True, print all debug output from sub-steps;
                            if False, only print summary lines (errors are still shown)
    """
    # -----------------------------------------------------------------------
    # Resolve all file paths: fill in any None values from env vars or defaults,
    # then switch from Drive paths to local paths if Drive is not available.
    # -----------------------------------------------------------------------
    if clients_file is None:
        clients_file = os.environ.get("CLIENTS_FILE")

    if country_codes_file is None:
        country_codes_file = os.environ.get("COUNTRY_CODES_FILE")

    if output_dir is None:
        output_dir = os.environ.get("OUTPUT_DIR")
    if output_dir is None:
        output_dir = HARDCODED_OUTPUT_DIR

    if not _drive_available():
        # Running on a local machine: switch all Drive paths to local Windows paths
        print("[*] Drive not available; running and saving on local machine.")
        if input_folder in (None, HARDCODED_INPUT_FOLDER):
            input_folder = LOCAL_INPUT_FOLDER
        if clients_file == HARDCODED_CLIENTS_FILE:
            _local_cl = Path(LOCAL_CLIENTS_FILE)
            clients_file = str(_local_cl) if _local_cl.is_file() else None
        if country_codes_file == HARDCODED_COUNTRY_CODES_FILE:
            _local_cc = Path(LOCAL_COUNTRY_CODES_FILE)
            country_codes_file = str(_local_cc) if _local_cc.is_file() else None
        if output_dir == HARDCODED_OUTPUT_DIR:
            output_dir = LOCAL_OUTPUT_DIR
        if archive_folder in (None, HARDCODED_ARCHIVE_FOLDER):
            archive_folder = LOCAL_ARCHIVE_FOLDER
    else:
        # Running in Colab with Drive mounted: use Drive paths where available,
        # fall back to local paths only for paths that don't exist on Drive
        input_folder = _use_drive_or_local(input_folder, LOCAL_INPUT_FOLDER, is_dir=True) if input_folder else input_folder
        if clients_file:
            clients_file = _use_drive_or_local(clients_file, LOCAL_CLIENTS_FILE)
        else:
            _dc_cl = Path(HARDCODED_CLIENTS_FILE)
            _lc_cl = Path(LOCAL_CLIENTS_FILE)
            _proj_cl = PROJECT_ROOT / "addition" / "clients.txt"
            if _dc_cl.is_file():
                clients_file = str(_dc_cl)
            elif _lc_cl.is_file():
                clients_file = str(_lc_cl)
            elif _proj_cl.is_file():
                clients_file = str(_proj_cl.resolve())
        if country_codes_file:
            country_codes_file = _use_drive_or_local(country_codes_file, LOCAL_COUNTRY_CODES_FILE)
        else:
            _dc_cc = Path(HARDCODED_COUNTRY_CODES_FILE)
            _lc_cc = Path(LOCAL_COUNTRY_CODES_FILE)
            _proj_cc = PROJECT_ROOT / "addition" / "dhl_country_codes.txt"
            if _dc_cc.is_file():
                country_codes_file = str(_dc_cc)
            elif _lc_cc.is_file():
                country_codes_file = str(_lc_cc)
            elif _proj_cc.is_file():
                country_codes_file = str(_proj_cc.resolve())
        output_dir = _use_drive_or_local(output_dir, LOCAL_OUTPUT_DIR, is_dir=True)
        if archive_folder:
            archive_folder = _use_drive_or_local(archive_folder, LOCAL_ARCHIVE_FOLDER, is_dir=True)

    # Create the output and processing folders if they don't already exist
    output_root = Path(output_dir) if output_dir else (PROJECT_ROOT / "output")
    output_root.mkdir(parents=True, exist_ok=True)

    # processing/ lives next to output/ — on Drive when in Colab, local otherwise.
    # Deriving it from output_root ensures it ends up on the same storage as the output.
    processing_root = output_root.parent / "processing"
    processing_root.mkdir(parents=True, exist_ok=True)

    # Build output file names based on the input file's stem (name without extension).
    # e.g. input "myfile.json" -> outputs "myfile.xlsx", Country Regions TXT, Postal Code Zones TXT
    input_stem = Path(input_file).stem

    def _unique_path(directory, base_stem, suffix):
        """
        Return a file path that does not already exist.
        If base_stem + suffix exists, try base_stem_1 + suffix, base_stem_2 + suffix, etc.
        This prevents overwriting previous outputs when the same input is processed again.
        """
        candidate = directory / f"{base_stem}{suffix}"
        if not candidate.exists():
            return candidate
        for i in range(1, 10000):
            candidate = directory / f"{base_stem}_{i}{suffix}"
            if not candidate.exists():
                return candidate
        raise RuntimeError(f"Could not find unique path for {base_stem}{suffix}")

    output_xlsx_path = _unique_path(output_root, input_stem, ".xlsx")
    output_txt_path = _unique_path(output_root, input_stem + "_CountryZoning_by_RateName", ".txt")
    output_postal_txt_path = _unique_path(output_root, input_stem + "_Postal_Code_Zones", ".txt")
    extracted_json_path = processing_root / "from_rate_card_excel.json"

    # Resolve path for read_client_list (missing file → empty list).
    default_clients = PROJECT_ROOT / "addition" / "clients.txt"
    clients_path = Path(clients_file) if clients_file else default_clients

    # Print a summary of what the pipeline is about to do
    print("=" * 70)
    print("DHL PIPELINE RUNNER")
    print("=" * 70)
    print(f"[*] Project root: {PROJECT_ROOT}")
    print(f"[*] Input: {input_file}")
    print(f"[*] Clients file: {clients_path}" + ("" if clients_path.is_file() else " (optional — not found, client list empty)"))
    if country_codes_file:
        print(f"[*] Country codes file: {country_codes_file}")
    print(f"[*] Output directory: {output_root}")
    print(f"[*] Processing directory: {processing_root}")
    if input_folder:
        print(f"[*] Input folder: {input_folder}")
    if archive_folder:
        print(f"[*] Archive folder: {archive_folder}")
    print()

    # -----------------------------------------------------------------------
    # DEBUG: list contents of the four key folders so it is easy to verify
    # which files are visible to the pipeline at runtime.
    # -----------------------------------------------------------------------
    def _list_folder(label, folder_path):
        """Print all files in a folder, or a clear message if it doesn't exist."""
        p = Path(folder_path) if folder_path else None
        print(f"[DEBUG] {label}: {p}")
        if p is None:
            print("        (not configured)")
            return
        if not p.exists():
            print("        (folder does not exist)")
            return
        if not p.is_dir():
            print("        (path exists but is not a folder)")
            return
        files = sorted(p.iterdir())
        if not files:
            print("        (folder is empty)")
        else:
            for f in files:
                size = f.stat().st_size if f.is_file() else 0
                kind = "DIR " if f.is_dir() else f"FILE {size:>10,} bytes"
                print(f"        {kind}  {f.name}")

    print("[DEBUG] ---- Folder contents at pipeline start ----")
    _list_folder("INPUT  folder", input_folder or PROJECT_ROOT / "input")
    _list_folder("ARCHIVE folder", archive_folder or PROJECT_ROOT / "archive")
    _list_folder("OUTPUT  folder", output_root)
    print("[DEBUG] ---- End folder listing ----")
    print()

    # Copy reference files to the locations where create_table.py expects them
    _prepare_reference_files(country_codes_file)

    def _run_quiet(label, fn, *args, **kwargs):
        """
        Run a function and suppress its print output unless verbose=True.

        In quiet mode, all stdout and stderr from the function are captured
        and hidden.  If the function raises an exception, the last 20 lines
        of captured output are printed to help diagnose the error.

        This keeps the pipeline output clean and readable for the user,
        while still showing full detail when something goes wrong.
        """
        if verbose:
            return fn(*args, **kwargs)   # verbose mode: let all output through
        out_buf = io.StringIO()    # buffer to capture stdout
        err_buf = io.StringIO()    # buffer to capture stderr
        try:
            with contextlib.redirect_stdout(out_buf), contextlib.redirect_stderr(err_buf):
                return fn(*args, **kwargs)
        except Exception:
            print(f"[ERROR] {label} failed.")
            captured = out_buf.getvalue().strip()
            captured_err = err_buf.getvalue().strip()
            if captured:
                print("---- Captured stdout (last lines) ----")
                print("\n".join(captured.splitlines()[-20:]))   # show last 20 lines
            if captured_err:
                print("---- Captured stderr (last lines) ----")
                print("\n".join(captured_err.splitlines()[-20:]))
            raise

    # -----------------------------------------------------------------------
    # Step 1: Read client list (for metadata client hint on workbooks).
    # -----------------------------------------------------------------------
    print("Step 1: Reading client list...")
    client_list = _run_quiet("Read client list", read_client_list, str(clients_path))
    print(f"[OK] Client names loaded: {len(client_list)}")
    print()

    # -----------------------------------------------------------------------
    # Step 2: Load Rate Card workbook or extracted JSON (conversion-to-json.py).
    # -----------------------------------------------------------------------
    print("Step 2: Loading data (conversion-to-json)...")
    processed_data, client_name = _run_quiet(
        "Load Rate Card / JSON",
        load_extracted_payload_from_input,
        input_file,
        client_list,
    )

    FileName = Path(input_file).name
    processed_data.setdefault("metadata", {})["FileName"] = FileName

    _run_quiet("Save extracted JSON", _save_extracted_json, processed_data, str(extracted_json_path))
    stats = processed_data.get("statistics", {})
    print(f"[OK] Client: {client_name}")
    print(
        f"[OK] Rows: MainCosts={stats.get('MainCosts_rows', 0)}, "
        f"CountryZoning={stats.get('CountryZoning_rows', 0)}, "
        f"AccessorialCosts2={stats.get('AccessorialCosts2_rows', 0)}"
    )
    print()
    # -----------------------------------------------------------------------
    # Step 3: Fill in any MainCosts sections that have a null service_type.
    # This can happen when the PDF layout doesn't repeat the service name on every page.
    # fill_null_service_types() propagates the last known service_type forward.
    # The updated data is saved back to the extracted JSON file.
    # -----------------------------------------------------------------------
    print("Step 3: Filling null service_type values...")
    filled_count = fill_service_types.fill_null_service_types(processed_data)
    with open(extracted_json_path, "w", encoding="utf-8") as f:
        json.dump(processed_data, f, indent=2, ensure_ascii=False)
    print(f"[OK] Filled {filled_count} section(s)")
    print()

    # -----------------------------------------------------------------------
    # Step 4: Build the Excel workbook from the extracted data dictionary.
    # -----------------------------------------------------------------------
    print("Step 4: Creating Excel workbook...")
    output_xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    excel_summary = _run_quiet(
        "Create Excel", create_table.save_to_excel, processed_data, str(output_xlsx_path)
    )
    print(f"[OK] Excel created: {output_xlsx_path}")
    acc = (excel_summary or {}).get("accessorial")
    if acc is not None:
        n = acc.get("rows", 0)
        ref = acc.get("reference_file")
        if acc.get("sheet_written"):
            if ref:
                print(
                    f"[OK] Accessorial Costs tab: {n} row(s); "
                    f"Cost Type list from {ref}"
                )
            else:
                print(
                    f"[OK] Accessorial Costs tab: {n} row(s); "
                    "no client-matching reference file (Cost Type may be empty)"
                )
        else:
            print(
                "[*] Accessorial builder: 0 rows (Accessorial Costs tab not added)"
            )
    print()

    # -----------------------------------------------------------------------
    # Step 4b: Save the full Excel to processing/, then trim the output copy
    # to only the tabs needed by downstream consumers (see KEEP_TABS below).
    # -----------------------------------------------------------------------
    print("Step 4b: Saving full workbook to processing/ and trimming output...")
    import shutil as _shutil
    import openpyxl as _openpyxl

    # Move full xlsx to processing/
    processing_xlsx_path = processing_root / output_xlsx_path.name
    _shutil.copy2(str(output_xlsx_path), str(processing_xlsx_path))
    print(f"[*] Full workbook copied to processing: {processing_xlsx_path}")

    # Rewrite the output xlsx keeping only the required tabs
    KEEP_TABS = [
        "Metadata",
        "MainCosts",
        "CountryZoning",
        "AdditionalZoning",
        "AccessorialCosts2",
        "Accessorial Costs",
        "GoGreenPlusCost",
        "DemandSurcharge",
    ]
    try:
        _wb = _openpyxl.load_workbook(str(output_xlsx_path))
        tabs_to_remove = [s for s in _wb.sheetnames if s not in KEEP_TABS]
        for tab in tabs_to_remove:
            del _wb[tab]
        _wb.save(str(output_xlsx_path))
        kept = [s for s in _openpyxl.load_workbook(str(output_xlsx_path), read_only=True).sheetnames]
        print(f"[OK] Output workbook trimmed to tabs: {kept}")
    except Exception as e:
        print(f"[WARN] Could not trim output workbook (non-fatal): {e}")
    print()

    # -----------------------------------------------------------------------
    # Step 5: Build the CountryZoning TXT file from the Excel workbook.
    # This reads the CountryZoning tab of the Excel file and writes a plain-text
    # summary: one line per RateName listing all country codes for that rate.
    # -----------------------------------------------------------------------
    print("Step 5: Creating CountryZoning TXT...")
    txt_out = _run_quiet(
        "Create CountryZoning TXT",
        create_country_region_txt,
        excel_path=str(output_xlsx_path),
        output_path=str(output_txt_path),
        extracted_json_path=str(extracted_json_path),
        postal_output_path=str(output_postal_txt_path),
    )
    print(f"[OK] Country Regions TXT saved: {txt_out}")
    print(f"[OK] Postal Code Zones TXT saved: {output_postal_txt_path}")
    print()

    # Print the final success summary
    print("=" * 70)
    print("[SUCCESS] PIPELINE COMPLETE")
    print("=" * 70)
    print(f"Client: {client_name}")
    print(f"Extracted JSON: {extracted_json_path}")
    print(f"Excel: {output_xlsx_path}")
    print(f"Country Regions TXT: {output_txt_path}")
    print(f"Postal Code Zones TXT: {output_postal_txt_path}")

    # Move the input JSON file to the archive folder now that processing is complete
    archived_to = _archive_processed_input(
        input_file=input_file,
        input_folder=input_folder,
        archive_folder=archive_folder,
    )
    if archived_to:
        print(f"Archived input JSON: {archived_to}")
    print()

    # Print a clean "Overall" summary for easy copy-pasting into logs or emails
    print("Overall:")
    print(f"- Input processed: {input_file}")
    print(f"- Client: {client_name}")
    print(f"- JSON output: {extracted_json_path}")
    print(f"- Excel output: {output_xlsx_path}")
    print(f"- Country Regions TXT: {output_txt_path}")
    print(f"- Postal Code Zones TXT: {output_postal_txt_path}")
    if archived_to:
        print(f"- Archived input: {archived_to}")
    print()


def main():
    """
    Entry point when the script is run from the command line.
    Parses arguments and calls run_pipeline() with the resolved values.
    """
    args = parse_args()
    # resolve_input_file handles the case where no --input-file was given
    # by showing an interactive file picker
    input_file, selected_folder = resolve_input_file(args.input_file, args.input_folder)
    run_pipeline(
        input_file=input_file,
        clients_file=args.clients_file,
        country_codes_file=args.country_codes_file,
        output_dir=args.output_dir,
        input_folder=selected_folder or args.input_folder,
        archive_folder=args.archive_folder,
        verbose=args.verbose,
    )


# Only run main() when this script is executed directly.
# Does NOT run when imported as a module by another script (e.g. in Colab).
if __name__ == "__main__":
    main()



