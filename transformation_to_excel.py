"""
Build a multi-tab Excel workbook from ``processing/from_rate_card_excel.json``.

Orchestrates: ``transform_main_costs``, ``transform_other_tabs``, ``excel_helpers``.
Includes an ``AccessorialCosts2`` tab when the JSON array is non-empty, and an
``Accessorial Costs`` tab (approved ``Cost Type`` mapping via ``accessorial_costs.py``)
when any accessorial source rows exist.
Run: ``python transformation_to_excel.py``
"""

import json
import os
from pathlib import Path

from transform_main_costs import (
    apply_zone_labels_to_main_costs,
    build_matrix_main_costs,
    expand_main_costs_lanes_by_zoning,
    sort_main_costs_rows_for_layout,
)
from transform_other_tabs import (
    build_demand_surcharge_excel_rows,
    build_zone_label_lookup,
    flatten_array_data,
    pivot_added_rates,
)
from excel_helpers import write_accessorial_sheet, write_matrix_sheet, write_sheet


def load_extracted_data(filepath):
    print(f"[*] Loading extracted data from: {filepath}")
    with open(filepath, 'r', encoding='utf-8') as f:
        data = json.load(f)
    print("[OK] Data loaded successfully")
    return data


def create_metadata_sheet(workbook, metadata):
    print("[*] Creating Metadata tab...")
    from openpyxl.styles import Font, PatternFill, Alignment

    ws = workbook.create_sheet("Metadata", 0)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)

    def _str(v):
        s = "" if v is None else (v.replace("\n", " ") if isinstance(v, str) else str(v))
        return s

    data = [
        ["Field", "Value"],
        ["Client", _str(metadata.get("client"))],
        ["Carrier", _str(metadata.get("carrier"))],
        ["Validity Date", _str(metadata.get("validity_date"))],
        ["FileName", _str(metadata.get("FileName"))],
        ["Extraction Date", _str(metadata.get("extraction_date"))],
        ["Extraction Source", _str(metadata.get("extraction_source"))],
    ]

    for row_idx, row_data in enumerate(data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 60
    print("[OK] Metadata tab created")


def save_to_excel(data, output_path):
    """Write workbook to ``output_path``.

    Returns a summary dict. When accessorial sources exist, includes
    ``accessorial``: ``rows``, ``reference_file``, ``sheet_written`` (from
    ``accessorial_costs.build_accessorial_costs_rows``) so callers can log
    mapping status without relying on captured stdout.
    """
    print(f"[*] Creating Excel file: {output_path}")

    try:
        import openpyxl
    except ImportError:
        print("[ERROR] openpyxl not installed!  pip install openpyxl")
        raise

    try:
        excel_build_summary: dict = {}
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        metadata = data.get('metadata', {})
        create_metadata_sheet(wb, metadata)

        main_costs_data = data.get('MainCosts', [])
        zoning_matrix = data.get('ZoningMatrix', [])
        country_zoning = data.get('CountryZoning', [])
        if main_costs_data:
            matrix_rows, category_specs = build_matrix_main_costs(
                main_costs_data, metadata, zoning_matrix, country_zoning
            )
            if zoning_matrix:
                matrix_rows = expand_main_costs_lanes_by_zoning(matrix_rows, zoning_matrix)
            if country_zoning:
                zone_label_lookup = build_zone_label_lookup(country_zoning)
                matrix_rows = apply_zone_labels_to_main_costs(matrix_rows, zone_label_lookup)
            matrix_rows = sort_main_costs_rows_for_layout(matrix_rows)
            write_matrix_sheet(wb, "MainCosts", matrix_rows, category_specs, metadata)

        added_rates = data.get('AddedRates', [])
        if added_rates:
            write_sheet(wb, "AddedRates", pivot_added_rates(added_rates, metadata), metadata)

        additional_costs_1 = data.get('AdditionalCostsPart1', [])
        if additional_costs_1:
            write_sheet(
                wb, "AdditionalCostsPart1",
                flatten_array_data(additional_costs_1, metadata, 'AdditionalCostsPart1'),
                metadata,
            )

        country_zoning = data.get('CountryZoning', [])
        if country_zoning:
            write_sheet(
                wb, "CountryZoning",
                flatten_array_data(country_zoning, metadata, 'CountryZoning'),
                metadata,
            )

        additional_zoning = data.get('AdditionalZoning', [])
        if additional_zoning:
            write_sheet(
                wb, "AdditionalZoning",
                flatten_array_data(additional_zoning, metadata, 'AdditionalZoning'),
                metadata,
            )

        gogreen_plus = data.get('GoGreenPlusCost', [])
        if gogreen_plus:
            write_sheet(
                wb, "GoGreenPlusCost",
                flatten_array_data(gogreen_plus, metadata, 'GoGreenPlusCost'),
                metadata,
            )

        zoning_matrix_rows = data.get('ZoningMatrix', [])
        if zoning_matrix_rows:
            write_sheet(
                wb, "ZoningMatrix",
                flatten_array_data(zoning_matrix_rows, metadata, 'ZoningMatrix'),
                metadata,
            )

        additional_costs_2 = data.get('AdditionalCostsPart2', [])
        if additional_costs_2:
            write_sheet(
                wb, "AdditionalCostsPart2",
                flatten_array_data(additional_costs_2, metadata, 'AdditionalCostsPart2'),
                metadata,
            )

        accessorial_costs_2 = data.get('AccessorialCosts2', [])
        if accessorial_costs_2:
            write_sheet(
                wb, "AccessorialCosts2",
                flatten_array_data(accessorial_costs_2, metadata, 'AccessorialCosts2'),
                metadata,
            )

        ac_part1 = data.get('AdditionalCostsPart1') or []
        ac_part2 = data.get('AdditionalCostsPart2') or []
        if ac_part1 or ac_part2 or accessorial_costs_2:
            from accessorial_costs import build_accessorial_costs_rows

            ac_rows, ref_used = build_accessorial_costs_rows(
                ac_part1,
                ac_part2,
                metadata,
                accessorial_costs_2_toolbox=accessorial_costs_2 or None,
            )
            excel_build_summary["accessorial"] = {
                "rows": len(ac_rows),
                "reference_file": str(ref_used) if ref_used else None,
                "sheet_written": bool(ac_rows),
            }
            if ac_rows:
                write_accessorial_sheet(wb, "Accessorial Costs", ac_rows)

        demand_surcharge = data.get('DemandSurcharge') or []
        demand_costs = data.get('DemandCosts') or []
        if demand_surcharge or demand_costs:
            dr = build_demand_surcharge_excel_rows(demand_surcharge, demand_costs, metadata)
            if dr:
                write_sheet(wb, "DemandSurcharge", dr, metadata)

        wb.save(output_path)

        # Postal Code Zones text (used by expand_main_costs_with_postal_zones); same source as the sheet.
        cz_rows = data.get("CountryZoning") or []
        if cz_rows:
            from country_region_txt_creation import build_postal_code_zones_txt_lines

            postal_text = "\n".join(build_postal_code_zones_txt_lines(cz_rows))
            postal_fp = Path(output_path).parent / "Postal_Code_Zones.txt"
            postal_fp.parent.mkdir(parents=True, exist_ok=True)
            postal_fp.write_text(postal_text, encoding="utf-8")
            print(f"[*] Wrote {postal_fp.name} for postal / MainCosts correlation")

        try:
            from expand_additional_zoning import expand_main_costs_with_additional_zoning
            expand_main_costs_with_additional_zoning(output_path)
        except Exception as e:
            print(f"[WARN] MainCosts post-processing failed (non-fatal): {e}")

        try:
            from expand_main_costs_postal_zones import expand_main_costs_with_postal_zones
            expand_main_costs_with_postal_zones(output_path)
        except Exception as e:
            print(f"[WARN] Postal zone MainCosts expansion failed (non-fatal): {e}")

        file_size_kb = os.path.getsize(output_path) / 1024
        print(f"[OK] Excel file saved successfully")
        print(f"  - Tabs: {len(wb.sheetnames)}")
        print(f"  - File size: {file_size_kb:.2f} KB")

        return excel_build_summary

    except Exception as e:
        print(f"[ERROR] Failed to save Excel: {e}")
        raise


def main():
    print("=" * 60)
    print("RATE CARD → EXCEL")
    print("=" * 60)
    input_file = 'processing/from_rate_card_excel.json'
    output_dir = 'output'
    output_file = os.path.join(output_dir, 'UPS_Rate_Cards.xlsx')

    Path(output_dir).mkdir(parents=True, exist_ok=True)
    print(f"[OK] Output directory: {output_dir}\n")
    data = load_extracted_data(input_file)
    print("Building workbook…\n")
    save_to_excel(data, output_file)
    print()
    print("=" * 60)
    print("[SUCCESS]")
    print(f"Output: {output_file}")
    print("=" * 60)


if __name__ == "__main__":
    main()
